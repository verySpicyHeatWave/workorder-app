import os
import re
from datetime import date

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from appfiles.library.excelfilestatus import ExcelFileStatus
from appfiles.library.logcomment import LogComment
from appfiles.library.taskitem import TaskItem
from appfiles.utils.appglobals import default_building, default_room, IN_PROGRESS_DIR
from appfiles.utils.appglobals import primary_user
from appfiles.utils.utils import make_string_filepath_friendly, string_to_date

############### PUBLIC STATIC METHODS #########################################

def get_full_wo_number_from_xlsx_cell(ws: Worksheet) -> str:
    """A static method that gets the work order number from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and either returns
    the work order number by aggregating the appropriate cells OR it will return the
    next pending work order number if the WO number is not found to be valid.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    resp: str = f"{str(ws['A7'].value)[:6]}{str(ws['B7'].value)}{str(ws['C7'].value)}"
    if not is_a_valid_wo_number(resp):
        return determine_pending_number()
    return resp


def get_wo_number_from_xlsx_cell(ws: Worksheet) -> str:
    """A static method that gets the work order number from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and either returns
    the work order number by aggregating the appropriate cells OR it will return the
    next pending work order number if the WO number is not found to be valid.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    resp: str = f"{str(ws['A7'].value)[:6]}"
    if not is_a_valid_wo_number(resp):
        return determine_pending_number()
    return resp


def get_related_wo_from_xlsx_cell(ws: Worksheet, default: str = "N/A") -> str:
    """A static method that gets the related work order number from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and an optional default value
    which defaults to "N/A", and it either returns the default value (whether provided or not)
    or it returns the related work order number.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    if str(ws['G10'].value) == "None":
        return default
    resp: str = str(ws['G10'].value)
    if not is_a_valid_wo_number(resp):
        return default
    return resp


def is_a_valid_wo_number(value: str) -> bool:
    """A static method which analyzes a string using a regex pattern and
    determines whether or not it is a valid workorder.
    """
    wo_number_re = re.compile(r"\b\d{6}(FG|VB|CS|EA|FD|SX|FY|CL|BL|TH|)(S|N|I|\b)", re.IGNORECASE)
    regex = wo_number_re.search(value)
    return regex is not None


def get_ncr_number_from_xlsx_cell(ws: Worksheet, ncr_required: bool) -> str:
    """A static method that gets the related ncr number from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and a boolean value which
    represents whether or not an NCR is required, and it either returns the value from the cell,
    "REQUIRED" if the value in the cell is invalid, or "N/A" if an NCR is not required.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    if not ncr_required:
        return "N/A"
    value = str(ws['B12'].value)
    if is_a_valid_ncr_number(value):
        return value
    return "REQUIRED"



def is_a_valid_ncr_number(value: str) -> bool:
    """A static method which analyzes a string using a regex pattern and
    determines whether or not it is a valid NCR number OR the word 'REQUIRED'.
    """
    ncr_number_re = re.compile(r"NCR\d{6}W", re.IGNORECASE)
    regex = ncr_number_re.search(value)
    return regex is not None or value.upper() == "REQUIRED"


def get_date_from_xlsx_cell(ws: Worksheet, cellno: str) -> date:
    """A static method that gets the date from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and it either
    returns a date object representing the date in the planned execution cell
    or it returns today's date.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    resp: date = date.today()
    dateval = ws[cellno].value
    if isinstance(dateval, date):
        return dateval
    resp = string_to_date(str(dateval))
    return resp


def get_creator_from_xlsx_cell(ws: Worksheet) -> str:
    """A static method that gets the creator's name from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and it either
    returns the string in the associated cell or, if it's blank, the primary user's name.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    resp = str(ws['B10'].value)
    if resp == "None":
        resp = primary_user.name
    return resp


def get_priority_from_xlsx_cell(ws: Worksheet) -> int:
    """A static method that gets the priority assignment from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and it either returns the priority
    integer in the priority cell (assuming the value is between 1 and 3 inclusive) or it returns
    3.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    resp: int = 3
    try:
        resp = int(ws['A10'].value)
    except (TypeError, ValueError):
        resp = 3
    finally:
        if resp < 1 or resp > 3:
            resp = 3
    return resp


def get_building_from_xlsx_cell(ws: Worksheet) -> int:
    """A static method that gets the building number from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and it either returns the building
    integer in the location cell or it returns the user-defined default value. It determines
    whether or not the building number is really a building number using a regex pattern.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    cells: list[str] = ['D8', 'F8', 'H8', 'J8']
    contents: str = 'D8'
    bldg_reg = re.compile(r"(b|bldg)\.{,1}\s{,1}\d{2,5}\b", re.IGNORECASE)
    num_reg = re.compile(r"\d{2,5}")
    found: bool = False

    for c in cells:
        if str(ws[c].value) == "None":
            continue
        contents = str(ws[c].value)
        bldg = bldg_reg.search(contents)
        if bldg is not None:
            found = True
            break

    if not found:
        return default_building
    num = num_reg.search(bldg.group()) #type:ignore
    return int(num.group()) #type:ignore


def get_room_from_xlsx_cell(ws: Worksheet) -> int:
    """A static method that gets the room number from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and it either returns the room
    integer in the location cell or it returns the user-defined default value. It determines
    whether or not the room number is really a room number using a regex pattern.\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    cells: list[str] = ['D8', 'F8', 'H8', 'J8']
    contents: str = 'D8'
    room_reg = re.compile(r"(r|rm|room)\.{,1}\s{,1}\d{1,3}(\b|,|\\|/)", re.IGNORECASE)
    num_reg = re.compile(r"\d{1,3}")
    found: bool = False

    for c in cells:
        if str(ws[c].value) == "None":
            continue
        contents = str(ws[c].value)
        room = room_reg.search(contents)
        if room is not None:
            found = True
            break

    if not found:
        return default_room
    num_str = num_reg.search(room.group()) #type:ignore
    num = int(num_str.group()) #type:ignore
    if num < 1:
        return default_room
    return num


def get_title_from_xlsx_cell(ws: Worksheet) -> str:
    """A static method that gets the title string from the Excel workorder.\n
    It takes in a Worksheet object from the openpyxl library and it either returns the
    title string itself (up to a maximum of 90 characters, and will truncate otherwise) or
    it will return the string "No title provided".\n
    Note:
    ---------
    This method should be used in conjunction with the input_file_status() method or the
    approved_file_status() method to ensure that a real, defined work order file is being
    analyzed.
    """
    resp = str(ws['D7'].value)
    if resp == "None":
        return "No title provided"
    if len(resp) > 60:
        resp = resp[:60]
    return make_string_filepath_friendly(resp)


def get_task_list_from_xlsx_cells(ws: Worksheet) -> list[TaskItem]:
    """A static method which builds a list of TaskItem objects from the Excel workorder.
    It takes in a Worksheet object from the openpyxl library and analyzes the values from
    rows 15 through 29 Cells A, B, and G. If the cells have legitimate values, it stores
    their strings in a TaskItem data object.\n
    """
    resp: list[TaskItem] = []
    for i in range(15, 30):
        a_cell: str = f"A{i}"
        b_cell: str = f"B{i}"
        g_cell: str = f"G{i}"
        if ws[b_cell].value is not None:
            num: int = int(ws[a_cell].value)
            smr: str = str(ws[b_cell].value)
            ref: str = str(ws[g_cell].value)
            if ref == "None" and i > 15:
                ref = "REFERENCE REQUIRED"
            resp.append(TaskItem(num, smr, ref, i))
    if len(resp) < 2:
        resp.append(TaskItem(10, "TASK DESCRIPTION REQUIRED", "REFERENCE REQUIRED", 16))
    return resp


def get_comments_from_xlsx_cells(ws: Worksheet) -> list[LogComment]:
    """A static method which builds a list of LogComment objects from the Excel workorder.
    It takes in a Worksheet object from the openpyxl library and analyzes the values from
    rows 86 through 109 Cells A, I, and K. If the cells have legitimate values, it stores
    their strings in a LogComment data object.\n
    """
    resp: list[LogComment] = []
    for i in range(24):
        row = i + 86
        a_cell: str = f"A{row}"
        i_cell: str = f"I{row}"
        k_cell: str = f"K{row}"
        if ws[a_cell].value is not None:
            text: str = str(ws[a_cell].value)
            name: str = str(ws[i_cell].value)
            c_date: date = get_date_from_xlsx_cell(ws, k_cell)
            resp.append(LogComment(text, name, c_date, i))
    return resp


def determine_pending_number() -> str:
    """Method which determines what the Pending work order number should be based on how many
    other pending work orders are in the in_progress directory. It fills in the lowest available
    value.\n
    For example, if the list of pending workorders includes Pending-001, Pending-002, and
    Pending-005, then the next available pending work order number would be Pending-003.
    """
    nums: set[int] = __get_enumerated_file_numbers_set()
    if len(nums) < 1:
        return "Pending-001"
    resp: str = ""
    for i in range(len(nums)):
        index: int = i + 1
        resp = __generate_pending_string(index + 1)
        if index not in nums:
            return __generate_pending_string(index)
    return resp


def extract_pending_twois_number(name: str) -> int:
    """Extracts and returns the number value of a Pending workorder where the\n
     name is \"Pending-004.twois\" and the number to extract is 4.
     """
    return int(name.split(".")[0].split("-")[1])


def input_file_status(filepath: str) -> ExcelFileStatus:
    """Takes a path to a file and returns an ExcelFileStatus enum member depending on
    the cause of rejection. This method just checks to make sure that it is a valid TWOIS
    file by analyzing whether the file exists, has a valid excel extension, and that certain
    cells are populated with expected strings. It does not check to see that the form is properly
    filled out or approved.\n
    Potential members of the enum are defined in the documentation for the ExcelFileStatus
    enum class.
    """
    if not os.path.isfile(filepath):
        return ExcelFileStatus.NOT_FOUND
    if not filepath.endswith('.xlsx') and not filepath.endswith('.xls'):
        return ExcelFileStatus.NOT_XLSX

    resp: ExcelFileStatus = ExcelFileStatus.IS_VALID
    wb: Workbook = load_workbook(filename=filepath)
    ws: Worksheet = wb.active
    if (ws["D1"].value != 'Technician Work Order Information Sheet'
                    or ws['A13'].value != 'Work Order Plans'):
        resp = ExcelFileStatus.NOT_TWOIS

    wb.close()
    return resp


def approved_file_status(filepath: str) -> ExcelFileStatus:
    """Takes a path to a file and returns an ExcelFileStatus enum member depending on
    the cause of rejection. This method does everything that input_file_status() does, as well as
    verifies some cells to ensure that the form is properly filled out and is, indeed, approved.\n
    Potential members of the enum are defined in the documentation for the ExcelFileStatus
    enum class.
    """
    resp: ExcelFileStatus = input_file_status(filepath)
    if resp != ExcelFileStatus.IS_VALID:
        return resp

    wb: Workbook = load_workbook(filename=filepath)
    ws: Worksheet = wb.active

    if not __has_at_least_one_taskitem(ws) or str(ws['D7'].value) == "None":
        resp = ExcelFileStatus.NOT_COMPLETE

    wo_number: str = get_wo_number_from_xlsx_cell(ws)
    if resp == ExcelFileStatus.IS_VALID and (not is_a_valid_wo_number(wo_number) or
                                                not __has_pc_approval_markings(ws)):
        resp = ExcelFileStatus.NOT_APPROVED

    wb.close()
    return resp



def is_pac_required_per_xlsx(ws: Worksheet) -> bool:
    """A static method which analyzes the contents of the "PAC Required" cell and
    determines whether or not a PAC is required. Returns true if a PAC is required
    and false if not.
    """
    value: str = str(ws['I10'].value).lower()
    return value.endswith("yes") or value.endswith("required")




def format_room_number(room: int, building: int) -> str:
    """Takes in the room and building number and returns a formatted string. Particularly
    useful when bldg is 1768 and rm is 6, it returns 'Rm. 6/7/8'."""
    if room == 6 and building == 1768:
        return "Rm. 6/7/8"
    return f"Rm. {room}"



############### PRIVATE STATIC METHODS ########################################

def __get_enumerated_file_numbers_set() -> set[int]:
    """Analyzes the open workorders and looks at which ones are pendin Extracts the int
    from the pending WO number and adds it to a set.\n
    This is a private method that should only be used by the determine_pending_number() method.
    """
    resp: set[int] = set()
    for name in os.listdir(IN_PROGRESS_DIR):
        if name.endswith(".twois") and name.startswith("Pending"):
            num: int = extract_pending_twois_number(name)
            resp.add(num)
    return resp


def __generate_pending_string(num: int) -> str:
    """Takes an integer and builds and returns an appropriately formatted string,
    where the number is 4 and the result is \"Pending-004\".\n
    This is a private method that should only be used by the determine_pending_number() method.
    """
    n: str = str(num).rjust(max(3, len(str(num))), '0') #leading zeroes, minimum of 3 digits
    return f"Pending-{n}"


def __has_pc_approval_markings(ws: Worksheet) -> bool:
    """Takes an openpyxl Worksheet object and checks to see whether cells C3, C4, and C5
    are populated. Returns true if they all are, as this would mean that the workorder has
    been approved by PC.
    """
    cells: list[str] = ['C3', 'C4', 'C5']
    for c in cells:
        if str(ws[c].value) == "None":
            return False
    return True


def __has_at_least_one_taskitem(ws: Worksheet) -> bool:
    """Takes an openpyxl Worksheet object and checks to see whether cells A16, B16, G16, and B15
    are populated. Returns true if they all are, as this would mean that at least one task item
    with a reference is on the workorder.
    """
    cells: list[str] = ['A16', 'B16', 'G16', 'B15']
    for c in cells:
        if str(ws[c].value) == "None":
            return False
    return True
