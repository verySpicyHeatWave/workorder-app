#pylint: skip-file

from datetime import date, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import os
import shutil
from appfiles.library.workorder import WorkOrder
from .appglobals import IN_PROGRESS_DIR, TESTFILE, TEMPLATE_DIR
from .utils import date_to_string


testfile = os.path.join(IN_PROGRESS_DIR,"testfile.xlsx")

def clear_dir(directory: str, prefix: str = "") -> None:
    """Static function meant to clear all files from a directory recursively, but will leave the root directory in tact."""
    if not os.path.isdir(directory):
        return
    try:
        files = os.listdir(directory)
        for file in files:
            fp: str = os.path.join(directory,file)
            if os.path.isdir(fp):
                clear_dir(fp)
                os.rmdir(fp)
            elif file.startswith(prefix) and os.path.isfile(fp):
                os.remove(fp)
    except OSError:
        print("Could not delete files!")
    if directory == IN_PROGRESS_DIR:
        shutil.copyfile(TESTFILE, testfile)


def insert_approval_stamps(ws: Worksheet) -> None:    
    ws['C3'] = date_to_string(date.today())
    ws['C4'] = 'S'
    ws['C5'] = 'FSH'

    

def insert_approval_markings_on_wo(test_wo: WorkOrder):
    wb: Workbook = load_workbook(test_wo.get_excel_filepath())
    insert_approval_stamps(wb.active)
    wb.save(test_wo.get_excel_filepath())
    wb.close()

def get_date_dir_name(d: date) -> str:
    MONDAY: int = 0
    one_day: timedelta = timedelta(days=1)
    while d.weekday() != MONDAY:
        d -= one_day
    return f'Week of {date_to_string(d, "%m-%d")}'



def write_to_single_cell_in_test_file(value: str, cellno: str) -> None:
    """Static function meant to write a specific string to a specific cell in the test file."""
    clear_dir(IN_PROGRESS_DIR)
    wb: Workbook = load_workbook(testfile)
    ws: Worksheet = wb.active
    ws[cellno].value = value
    wb.save(testfile)
    wb.close()
