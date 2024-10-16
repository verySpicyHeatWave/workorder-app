"""Module which provides the WorkOrder class."""


############### IMPORT STATEMENTS #############################################

import os
import pickle
from datetime import date
from typing import Self, Unpack

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from appfiles.library.completiondata import WorkorderCompletionData
from appfiles.utils.appglobals import COMPLETE_DIR, IN_PROGRESS_DIR, TEMPLATE_DIR, TEMPLATE_TWOIS
from appfiles.utils.appglobals import default_building, default_room
from appfiles.utils.appglobals import primary_user
from appfiles.utils.utils import bool_to_yes_no_string, name_to_initials
from appfiles.utils.utils import create_dated_directories, yes_no_string_to_bool, is_within_bounds
from appfiles.utils.utils import safe_rename, date_to_string, make_string_filepath_friendly
import appfiles.utils.workorder as woutils

from appfiles.library.excelfilestatus import ExcelFileStatus
from appfiles.library.logcomment import LogComment
from appfiles.library.site import Site, default_site
from appfiles.library.special import Special, default_special
from appfiles.library.taskitem import TaskItem
from appfiles.library.workorder_dict import WorkOrderDict
from appfiles.library.workorder_type import WorkOrderType, default_wotype

SAFETY_TASK: TaskItem = TaskItem(0, "Safety Message", "", 15)
REQUIRED_TASK: TaskItem = TaskItem(10, "TASK DESCRIPTION REQUIRED", "REFERENCE REQUIRED", 16)


############### PRIMARY CLASS DEFINITION ######################################

class WorkOrder:
    """
    Overview
    -------------------------
    A class representing a TWOIS work order.\n
    The object will use the other parameters fed to it as
    arguments and, if parameters are not supplied, then defaults will
    be used instead. All of the parameters are optional keyword arguments.

    Fields: Names, Types, and Default Values
    -------------------------
        description : str
            An optional description of the task
            Default: "No description of task provided"

        due_date : date
            The date that the work is scheduled to be performed.
            Default: Today's date

        wo_number : str
            The string that represents the workorder. Is either \"Pending-###\"
            or a six-digit approved TWOIS number.
            Default: The next available "Pending" workorder number.

        site : Site(Enum) = 
            An enum that represents all of the different sites available on the
            TWOIS form.
            Default: Defined by user in global configuration

        special : Special(Enum)
            Defines the department carrying out the work. N is or Northrop 
            Grumman, I is for ISSOs, and S is for SysAdmin.
            Default: Defined by user in global configuration

        title : str
            The title of the workorder. Typically a brief description 
            (like \"14-Day AV Updates (OSB GMM)\")
            Default: "No title provided"

        type : WorkOrderType(Enum)
            An enum that represents the type of work order. Preventative
            Maintenance, Corrective Maintenance, or Other are the available
            options.
            Default: Defined by user in global configuration

        priority : int
            A priority assignment to determine level of importance. Must be
            between 1 and 3.
            Default: 3

        creator : str
            The name of the creator of this workorder. The default is almost
            always sufficient, as it is a system-level configuration.
            Default: Defined by user in global configuration (Primary User field)

        building : int
            The building number that the work will take place in.
            Default: Defined by user in global configuration

        room : int
            The room number that the work will take place in.
            Default: Defined by user in global configuration

        related_wo : str
            An optional value that links this work order to a prior work
            order (e.g. if this is a corrective maintenance item following
            a discovery from a prior preventative maintenance item)
            Default: "N/A"

        pac_required : bool
            Indicates whether or not a PAC is required.
            Default: False

        ncr_required : bool
            Indicates whether or not an NCR is required.    
            Default: False

        ncr_number : str
            If an NCR is required, this field holds the NCR number. This
            field is optional if ncr_required is False.
            Default: "N/A"

        task_lead_required : bool
            Indicates whether a task lead's presence is required.
            Default: False

        tech_witness_point : bool
            Indicates whether a technical witness's presence is required.
            Default: False

        peer_review_required : bool
            Indicates whether a peer review is required.
            Default: False
            
        peer_review_attached : bool
            Indicates whether a peer review is attached to this work order.
            Default: False

        ehs_required : bool
            Indicates whether EHS's presence is required.
            Default: False  

        qamip : bool
            Indicates... QAMIP?
            Default: False

        qa_review_required : bool
            Indicates whether a review of the work by QA is required.
            Default: False

        task_list : list[TaskItem]
            A list of all of the tasks to be carried out. Must be in the
            form of a list[TaskItem].
            Default: A list[TaskItem] with a single TaskItem defined as
            follows:
                number = 0
                summary = "Safety Message"
                reference = ""
                planned_row = 15
                actuals_row = 32
            Note: For more details on the TaskItem object, check the documentation
            associated with the TaskItem class.

        comments: list[LogComments]
            A list of all of the log comments. Must be in the
            form of a list[LogComments].
            Default: An empty list[LogComments]
            Note: For more details on the LogComment object, check the documentation
            associated with the LogComment class.
        
        completion_data: CompletionData | None
            An optional field which is equal to None, and should remain so until
            the workorder is completed. It consists of a start date, end date,
            restoration date, and repair time.

    Public Methods        
    -------------------------
        save() -> None:
            Method which saves the workorder both as a .twois file, which is a pickle serialization
            of the object, as well as saves the .xlsx workorder spreadsheet. An optional filepath
            can be provided to determine the location of the file. If left blank, the files will be
            saved into the in_progress directory.

        edit(**kwargs: Unpack[WorkOrderDict]) -> None:
            A method which enables a user to edit an existing WorkOrder object by entering a new
            value for any of its fields using kwargs.

        approve(filename: str) -> ExcelFileStatus:
            A method which "approves" the WorkOrder object. It takes in a path to a file and
            analyzes it, and if it determines that the workorder is approved, it delets the
            existing work order files and replaces them with the approved files. This should
            only be called on a pending workorder.

        delete() -> None:
            A method which is used to permanently delete a work order's files.

        save_as_wo_template(filename: str) -> str:
            Method which serializes the WorkOrder object as a .twois file using the
            pickle library. This file is typically saved in the 'templates' directory. It takes in a
            name for the file. If you put a full existsing filepath or a filename including the
            .twois in as the argument, the workorder will simply be created, overwriting the
            existing file.

        get_full_workorder_number() -> str:
            A simple method which returns the full workorder number, which is formatted as
            123456VBS, where 123456 are the workorder number proper, VB is the site, and S is the
            department signifier. If the workorder is not approved, will just return the 'Pending'
            workorder number.

        log_comment(comments: list[LogComment]) -> bool:
            Method which takes in a list of LogComments and appends them to this object's comments.
            Returns true if the operation was successful, and false if the list is full. The list
            maxes out at 24 logs.

        complete(completion_date: date) -> None:
            Method that is called when a workorder is complete. Takes in a completion date which
            should be gotten from the form that will call this method. It deletes the .twois
            file and then moves the excel file into the "complete/Week of MM/DD"
            directory, and generates a text file containing the description field to accompany
            the excel spreadsheet.

        matches_file(filepath: str) -> bool:
            A method which takes in a path to a file. It returns true if the file is a legitimate
            excel TWOIS file that matches the details of this workorder. Otherwise, it returns 
            false.

        is_approved() -> bool:
            Analyzes the work order number which is either "Pending-###" or "######ABC" and
            determines whether or not a workorder is approved. If the workorder has been assigned
            a valid workorder number, it returns true. Otherwise, it returns false.

        get_excel_filepath() -> str:
        Returns the excel filepath associated with this work order.

        get_twois_filepath() -> str:
        Returns the twois filepath associated with this work order.
    """

    #BCOBB: Verify that everything above is still accurate once this is done.

    ########### CLASS CONSTANTS
    MAX_DESCRIPTION: int = 280
    DEFAULT_DESCRIPTION: str = "No description of task provided"
    MAX_TITLE: int = 60
    DEFAULT_TITLE: str = "No title provided"


    ########### DUNDER METHODS
    def __init__(self, **kwargs: Unpack[WorkOrderDict]) -> None: #type:ignore
        # Declare fields and set defaults
        self.description: str = self.DEFAULT_DESCRIPTION
        self.due_date: date = date.today()
        self.wo_number: str = woutils.determine_pending_number()
        self.site: Site = default_site
        self.special: Special = default_special
        self.title: str = self.DEFAULT_TITLE
        self.type: WorkOrderType = default_wotype
        self.priority: int = 3
        self.creator: str = primary_user.name
        self.building: int = default_building
        self.room: int = default_room
        self.related_wo: str = "N/A"
        self.pac_required: bool = False
        self.ncr_required: bool = False
        self.ncr_number: str = "N/A"
        self.task_lead_required: bool = False
        self.tech_witness_point: bool = False
        self.peer_review_required: bool = False
        self.peer_review_attached: bool = False
        self.ehs_required: bool = False
        self.qamip: bool = False
        self.qa_review_required: bool = False
        self.task_list: list[TaskItem] = [SAFETY_TASK, REQUIRED_TASK]
        self.comments: list[LogComment] = []
        self.completion_data: WorkorderCompletionData | None = None

        # Populate the fields with kwargs and establish private filenames
        self.__populate_fields(**kwargs)
        self.__save_dir: str = IN_PROGRESS_DIR
        self.__twois_filename: str = f"{self.get_full_workorder_number()}.twois"
        self.__excel_filename: str = f"{self.get_full_workorder_number()} - {self.title}.xlsx"


    def __str__(self) -> str:
        return f"{self.wo_number}: {self.title}"


    def __eq__(self, other: 'WorkOrder') -> bool: #type:ignore
        if self.is_approved():
            return self.wo_number == other.wo_number
        return (self.title != other.title and
                self.task_list[1].summary != other.task_list[1].summary and
                self.due_date != other.due_date)


    ########### PRIVATE METHODS
    def __populate_fields(self, **kwargs: Unpack[WorkOrderDict]) -> None: #type:ignore
        kw = kwargs.keys()
        if 'description' in kw and kwargs['description'].isalnum():
            self.description = kwargs['description'][:self.MAX_DESCRIPTION]

        if 'due_date' in kw:
            self.due_date = kwargs['due_date']

        if 'site' in kw:
            self.site = kwargs['site']

        if 'special' in kw:
            self.special = kwargs['special']

        if 'wo_number' in kw and woutils.is_a_valid_wo_number(kwargs['wo_number']):
            self.wo_number = f"{kwargs['wo_number'][:6]}"

        if 'title' in kw:
            self.title = make_string_filepath_friendly(kwargs['title'][:self.MAX_TITLE])

        if 'type' in kw:
            self.type = kwargs['type']

        if 'priority' in kw and is_within_bounds(kwargs['priority'], 1, 3):
            self.priority = kwargs['priority']

        if 'creator' in kw:
            self.creator = kwargs['creator']

        if 'building' in kw and is_within_bounds(kwargs['building'], 10, 99_999):
            self.building = kwargs['building']

        if 'room' in kw and is_within_bounds(kwargs['room'], 1, 999):
            self.room = kwargs['room']

        if 'related_wo' in kw and woutils.is_a_valid_wo_number(kwargs['related_wo']):
            self.related_wo = kwargs['related_wo']

        if 'pac_required' in kw:
            self.pac_required = kwargs['pac_required']

        if 'ncr_required' in kw:
            self.ncr_required = kwargs['ncr_required']
            self.ncr_number = 'REQUIRED' if self.ncr_required else 'N/A'

        if ('ncr_number' in kw and self.ncr_required and
                    woutils.is_a_valid_ncr_number(kwargs['ncr_number'])):
            self.ncr_number = kwargs['ncr_number']

        if 'task_lead_required' in kw:
            self.task_lead_required = kwargs['task_lead_required']

        if 'tech_witness_point' in kw:
            self.tech_witness_point = kwargs['tech_witness_point']

        if 'peer_review_required' in kw:
            self.peer_review_required = kwargs['peer_review_required']

        if 'peer_review_attached' in kw:
            self.peer_review_attached = kwargs['peer_review_attached']

        if 'ehs_required' in kw:
            self.ehs_required = kwargs['ehs_required']

        if 'qamip' in kw:
            self.qamip = kwargs['qamip']

        if 'qa_review_required' in kw:
            self.qa_review_required = kwargs['qa_review_required']

        if 'task_list' in kw:
            self.task_list = kwargs['task_list']

        if 'comments' in kw:
            self.comments = kwargs['comments']


    @staticmethod
    def __get_kwargs_from_xlsx(filepath: str, descrip: str) -> WorkOrderDict:
        wb: Workbook = load_workbook(filepath)
        ws: Worksheet = wb.active

        _ncr_required: bool = yes_no_string_to_bool(str(ws['A12'].value))

        resp: WorkOrderDict = WorkOrderDict(
            due_date = woutils.get_date_from_xlsx_cell(ws, 'B3'),
            wo_number = woutils.get_wo_number_from_xlsx_cell(ws),
            site = Site.parse(str(ws['B7'].value)),
            special = Special.parse(str(ws['C7'].value)),
            title = woutils.get_title_from_xlsx_cell(ws),
            type = WorkOrderType.parse(str(ws['I7'].value)),
            priority = woutils.get_priority_from_xlsx_cell(ws),
            creator = woutils.get_creator_from_xlsx_cell(ws),
            building = woutils.get_building_from_xlsx_cell(ws),
            room = woutils.get_room_from_xlsx_cell(ws),
            related_wo = woutils.get_related_wo_from_xlsx_cell(ws),
            pac_required = woutils.is_pac_required_per_xlsx(ws),
            ncr_required = _ncr_required,
            ncr_number = woutils.get_ncr_number_from_xlsx_cell(ws, _ncr_required),
            task_lead_required = yes_no_string_to_bool(str(ws['E12'].value)),
            tech_witness_point = yes_no_string_to_bool(str(ws['F12'].value)),
            peer_review_required = yes_no_string_to_bool(str(ws['G12'].value)),
            peer_review_attached = yes_no_string_to_bool(str(ws['H12'].value)),
            ehs_required = yes_no_string_to_bool(str(ws['I12'].value)),
            qamip = yes_no_string_to_bool(str(ws['J12'].value)),
            qa_review_required = yes_no_string_to_bool(str(ws['K12'].value)),
            task_list = woutils.get_task_list_from_xlsx_cells(ws),
            comments = woutils.get_comments_from_xlsx_cells(ws),
            description=descrip)
        return resp


    def __save_as_twois_file(self) -> str:
        """Method which serializes the WorkOrder object as a .twois file using the
        pickle library.
        """
        fp = f"{self.__save_dir}\\{self.__twois_filename}"
        with open(fp, "wb") as outfile:
            pickle.dump(self, outfile)
        return fp


    def __save_as_excel_file(self, load_file: str = TEMPLATE_TWOIS) -> str:
        """Method which saves the WorkOrder's data into a properly formatted .xlsx
        file using the openpyxl library. Takes in an optional filename, so that it could start
        by using an existing excel spreadsheet. If none is provided, it will use a blank template.
        """
        if (woutils.input_file_status(load_file) != ExcelFileStatus.IS_VALID or
                                                not self.matches_file(load_file)):
            load_file = TEMPLATE_TWOIS
        wb: Workbook = load_workbook(load_file)
        ws: Worksheet = wb.active

        ws['B3'].value = date_to_string(self.due_date)
        ws['B3'].alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')
        ws['B4'].value = 'RS'
        ws['B5'].value = name_to_initials(self.creator)
        ws['A7'].value = "" if self.wo_number.startswith("Pending") else self.wo_number[:6]
        ws['B7'].value = self.site.name
        ws['C7'].value = self.special.name
        ws['D7'].value = self.title
        ws['I7'].value = self.type.name
        ws['D8'].value = f"B{self.building}"
        ws['F8'].value = woutils.format_room_number(self.room, self.building)
        ws['A10'].value = f"{self.priority}"
        ws['B10'].value = self.creator
        ws['G10'].value = self.related_wo
        ws['I10'].value = "PAC - YES" if self.pac_required else "PAC - NO"
        ws['A12'].value = bool_to_yes_no_string(self.ncr_required)
        ws['B12'].value = self.ncr_number if self.ncr_required else "N/A"
        ws['E12'].value = bool_to_yes_no_string(self.task_lead_required)
        ws['F12'].value = bool_to_yes_no_string(self.tech_witness_point)
        ws['G12'].value = bool_to_yes_no_string(self.peer_review_required)
        ws['H12'].value = bool_to_yes_no_string(self.peer_review_attached)
        ws['I12'].value = bool_to_yes_no_string(self.ehs_required)
        ws['J12'].value = bool_to_yes_no_string(self.qamip)
        ws['K12'].value = bool_to_yes_no_string(self.qa_review_required)

        for i, task in enumerate(self.task_list):
            if i > 15:
                break
            task.planned_row = i + 15
            task.actuals_row = i + 32
            ws[f'A{str(task.planned_row)}'].value = task.number
            ws[f'B{str(task.planned_row)}'].value = task.summary
            ws[f'G{str(task.planned_row)}'].value = task.reference
            if task.is_complete():
                ws[f'A{str(task.actuals_row)}'].value = task.number
                ws[f'C{str(task.actuals_row)}'].value = date_to_string(task.completion_date)
                ws[f'E{str(task.actuals_row)}'].value = task.technician
                ws[f'H{str(task.actuals_row)}'].value = task.qty_techs
                ws[f'J{str(task.actuals_row)}'].value = f"{task.hours:0.1f}"

        for i, comment in enumerate(self.comments):
            if i > 24:
                break
            comment.set_row(i + 86)
            ws[f'A{str(comment.get_row())}'].value = comment.text
            ws[f'I{str(comment.get_row())}'].value = str(comment.person)
            ws[f'K{str(comment.get_row())}'].value = date_to_string(comment.date)

        if self.completion_data is not None:
            ws['A78'].value = date_to_string(self.completion_data.startdate)
            ws['C78'].value = date_to_string(self.completion_data.enddate)
            ws['A80'].value = date_to_string(self.completion_data.restoredate)
            ws['C80'].value = f"{str(self.completion_data)}"

        fp: str = f'{self.__save_dir}\\{self.__excel_filename}'
        wb.save(fp)
        wb.close()
        return fp


    ########### PUBLIC METHODS
    @classmethod
    def from_xlsx(cls, filename: str, description: str = DEFAULT_DESCRIPTION) -> Self:
        """A factory method which takes in the name of a file and determines whether or not
        it is a valid Work Order file. If it is, it will return a work order object based on
        the contents of that file. If it isn't, it will return a 'default' WorkOrder object.
        Also takes in an optional task description, since that is not part of the excel file.
        Typically, the description would be passed in either from the UI or from a WorkOrder
        object that is being approved.
        """
        if woutils.input_file_status(filename) == ExcelFileStatus.IS_VALID:
            resp: Self = cls(**WorkOrder.__get_kwargs_from_xlsx(filename, description))
            return resp
        return cls()


    @classmethod
    def from_twois(cls, filename: str) -> Self:
        """A factory method which takes in the name of a file and uses the pickle library to
        load that serialized object into memory. File must be a .twois file.
        """
        if filename.endswith(".twois") and os.path.isfile(filename):
            with open(filename, "rb") as infile:
                resp: Self = pickle.load(infile)
            return resp
        return cls()


    def save(self) -> None:
        """Method which saves the workorder both as a .twois file, which is a pickle serialization
        of the object, as well as saves the .xlsx workorder spreadsheet. An optional filepath
        can be provided to determine the location of the file. If left blank, the files will be
        saved into the in_progress directory.
        """
        self.__save_as_excel_file(f"{self.__save_dir}\\{self.__excel_filename}")
        self.__save_as_twois_file()


    def edit(self, **kwargs: Unpack[WorkOrderDict]) -> None: #type:ignore
        """A method which enables a user to edit an existing WorkOrder object by entering a new
        value for any of its fields using kwargs.\n
        See the documentation for the WorkOrder class for a description of those fields.
        """
        previous_excel: str = self.get_excel_filepath()
        previous_twois: str = self.get_twois_filepath()
        self.__populate_fields(**kwargs)
        self.__twois_filename = f"{self.get_full_workorder_number()}.twois"
        self.__excel_filename = f"{self.get_full_workorder_number()} - {self.title}.xlsx"
        safe_rename(previous_excel, self.get_excel_filepath())
        safe_rename(previous_twois, self.get_twois_filepath())
        self.save()


    def approve(self, filename: str, override: bool = False) -> ExcelFileStatus:
        """A method which "approves" the WorkOrder object. It takes in a path to a file and
        analyzes it, and if it determines that the workorder is approved, it delets the
        existing work order files and replaces them with the approved files. This should
        only be called on a pending workorder.
        """
        fstatus: ExcelFileStatus = ExcelFileStatus.IS_VALID
        if not override:
            if self.is_approved():
                return ExcelFileStatus.WO_ALREADY_APPROVED

            fstatus = woutils.approved_file_status(filename)
            if fstatus != ExcelFileStatus.IS_VALID:
                return fstatus

            if not self.matches_file(filename, True):
                return ExcelFileStatus.FILES_UNMATCHED

        new_workorder: WorkOrder = WorkOrder.from_xlsx(filename, self.description)
        self.delete()
        new_workorder.save()
        return fstatus


    def delete(self) -> None:
        """A method which is used to permanently delete a work order's files."""
        excel_file = f"{self.__save_dir}\\{self.__excel_filename}"
        twois_file = f"{self.__save_dir}\\{self.__twois_filename}"
        if os.path.exists(excel_file):
            os.remove(excel_file)
        if os.path.exists(twois_file):
            os.remove(twois_file)


    def save_as_wo_template(self, filename: str) -> str:
        """Method which serializes the WorkOrder object as a .twois file using the
        pickle library. This file is typically saved in the 'templates' directory. It takes in a
        name for the file. If you put a full existsing filepath or a filename including the .twois
        in as the argument, the workorder will simply be created, overwriting the existing
        file.
        """
        if filename.endswith(".twois"):
            if os.path.exists(filename):
                with open(filename, "wb") as outfile:
                    pickle.dump(self, outfile)
                return filename
            elif len(filename.split("\\")) > 1:
                filedir: str = "\\".join(filename.split("\\")[:-1])
                if os.path.isdir(filedir):
                    with open(filename, "wb") as outfile:
                        pickle.dump(self, outfile)
                    return filename
            else:
                filename = filename[:-6]

        new_fn: str = f"{TEMPLATE_DIR}\\{filename}-1.twois"
        i: int = 2
        while os.path.exists(new_fn) and os.path.isfile(new_fn) and i < 999:
            new_fn = f"{TEMPLATE_DIR}\\{filename}-{i}.twois"
            i += 1

        with open(new_fn, "wb") as outfile:
            pickle.dump(self, outfile)
        return new_fn


    def get_full_workorder_number(self) -> str:
        """A simple method which returns the full workorder number, which is formatted as
        123456VBS, where 123456 are the workorder number proper, VB is the site, and S is the
        department signifier. If the workorder is not approved, will just return the 'Pending'
        workorder number.
        """
        if self.is_approved(False):
            return f"{self.wo_number}{self.site.name}{self.special.name}"
        return self.wo_number


    def log_comment(self, comments: list[LogComment]) -> bool:
        """Method which takes in a list of LogComments and appends them to this object's comments.
        Returns true if the operation was successful, and false if the list is full. The list maxes
        out at 24 logs.
        """
        success: bool = True
        for new_comment in comments:
            row_is_taken: bool = False
            if len(self.comments) >= 24:
                success = False
                break
            for existing_comment in self.comments:
                if existing_comment.get_row() == new_comment.get_row():
                    row_is_taken = True
                    break
            if row_is_taken:
                success = False
                continue
            new_comment.set_row(len(self.comments) + 86)
            self.comments.append(new_comment)
        self.save()
        return success


    def complete(self, completion_data: WorkorderCompletionData) -> bool:
        """Method that is called when a workorder is complete. Takes in a completion date which
        should be gotten from the form that will call this method. It deletes the .twois
        file and then moves the excel file into the "complete/Week of MM/DD"
        directory, and generates a text file containing the description field to accompany
        the excel spreadsheet.
        """
        if not self.is_approved(False): #   BCOBB: REMOVE THE "FALSE" ARG--THIS IS FOR TEST ONLY
            return False

        tasks_are_complete: bool = True
        for task in self.task_list:
            if not task.is_complete():
                tasks_are_complete = False
                break

        if not tasks_are_complete:
            return False

        self.completion_data = completion_data
        old_dir: str = self.__save_dir
        self.__save_dir = create_dated_directories(COMPLETE_DIR, completion_data.enddate)
        self.__save_as_excel_file(f"{old_dir}\\{self.__excel_filename}")

        print("BCOBB: YOU HAVEN'T ADDED THE COMPLETION DATA TO THOSE CELLS YET!")

        with open(f"{self.__save_dir}\\{self.get_full_workorder_number()}_description.txt",
                  'x', encoding='utf-8') as file:
            file.write(self.description)
        self.__save_dir = old_dir
        self.delete()
        return True


    def matches_file(self, filepath: str, preapproved: bool = False) -> bool:
        """A method which takes in a path to a file. It returns true if the file is a legitimate
        excel TWOIS file that matches the details of this workorder. Otherwise, it returns false."""
        fstatus: ExcelFileStatus = ExcelFileStatus.IS_VALID
        if not preapproved:
            fstatus = woutils.approved_file_status(filepath)
            if fstatus != ExcelFileStatus.IS_VALID and fstatus != ExcelFileStatus.NOT_APPROVED:
                return False

        wb: Workbook = load_workbook(filepath)
        ws: Worksheet = wb.active
        resp: bool = False
        if fstatus == ExcelFileStatus.IS_VALID and self.is_approved():
            resp = self.get_full_workorder_number() == woutils.get_full_wo_number_from_xlsx_cell(ws)
        else:
            alltasksaregood: bool = True
            for i, thistask in enumerate(self.task_list):
                othertasks: list[TaskItem] = woutils.get_task_list_from_xlsx_cells(ws)
                if othertasks[i].summary != thistask.summary:
                    alltasksaregood = False
                    break
            resp = (self.due_date == woutils.get_date_from_xlsx_cell(ws, 'B3') and
                    self.title == woutils.get_title_from_xlsx_cell(ws) and
                    alltasksaregood)
        wb.close()
        return resp


    def is_approved(self, check_xlsx_file: bool = True) -> bool:
        """Analyzes the work order number which is either "Pending-###" or "######ABC" and
        determines whether or not a workorder is approved. If the workorder has been assigned
        a valid workorder number, it returns true. Otherwise, it returns false.
        """
        file_status: ExcelFileStatus = ExcelFileStatus.IS_VALID
        if check_xlsx_file:
            file_status = woutils.approved_file_status(self.get_excel_filepath())
        return (file_status == ExcelFileStatus.IS_VALID and
            woutils.is_a_valid_wo_number(f"{self.wo_number}{self.site.name}{self.special.name}"))


    def get_excel_filepath(self) -> str:
        """Returns the excel filepath associated with this work order."""
        return f"{self.__save_dir}\\{self.__excel_filename}"


    def get_twois_filepath(self) -> str:
        """Returns the twois filepath associated with this work order."""
        return f"{self.__save_dir}\\{self.__twois_filename}"



############### END OF FILE ###################################################
