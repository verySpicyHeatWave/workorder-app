################ IMPORT DECLARATIONS ########################################################################
import pickle, re

from datetime import date, datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

from lib.task_item import TaskItem
from lib.global_resources import *

import xlrd     #type:ignore


################ CLASS DEFINITION AND INITIALIZER ###########################################################
class WorkOrder:
    """A class representing a TWOIS work order. Must be generated from an existing Excel spreadsheet.\n
    Initializer takes in a file path to the existing Work Order excel spreadsheet and an optional description of the task defined by the Work Order."""
    def __init__(this, filepath: str, description: str = "No description of task provided") -> None:
        this.filepath: str = filepath
        this.description: str = description

        wb = load_workbook(this.filepath)
        ws = wb.active

        this.title: str = ws['D7'].value
        this.number: str = f"{str(ws['A7'].value)}{ws['B7'].value}{ws['C7'].value}" if str(ws['A7'].value) != "None" else get_next_pending_workorder_number()
        dateval = ws['B3'].value
        this.planned_date: date
        if isinstance(dateval, date):
            this.planned_date = dateval
        elif isinstance(dateval, str):
            ds: list[str] = ws['B3'].value.split('/')
            this.planned_date = date(int(ds[2]), int(ds[0]), int(ds[1]))
        this.workorder_type_index: int = int(this._determine_workorder_type_index(ws['I7'].value))
        this.priority_index: int = int(ws['A10'].value) - 1
        this.location_index: int = this._determine_location_index(ws['B7'].value, ws['D8'].value)
        this.site_string: str = ws['B7'].value
        this.location_string: str = ws['D8'].value

        noPAC: bool = str(ws['I10'].value).endswith("No") or ws['I10'].value == ""        
        this.requires_pac: bool = False if noPAC else True
        this.requires_ncr: bool = True if ws['A12'] == "Yes" else False
        this.requires_task_lead: bool = True if ws['E12'] == "Yes" else False
        this.tech_witness_point: bool = True if ws['F12'] == "Yes" else False
        this.requires_peer_review: bool = True if ws['G12'] == "Yes" else False
        this.peer_review_attached: bool = True if ws['H12'] == "Yes" else False
        this.requires_ehs: bool = True if ws['I12'] == "Yes" else False
        this.qamip: bool = True if ws['J12'] == "Yes" else False
        this.requires_qa_review: bool = True if ws['K12'] == "Yes" else False

        this.ncr_string: str = "N/A" if ws['A12'] == "No" else ws['B12'].value

        this.tasks: list[TaskItem] = this._get_task_list_from_excel_sheet(ws)



    ############ PUBLICALLY ACCESSIBLE METHODS ############################################################## 
    def save_workorder_as_excel_file(this) -> None:
        """Method which saves the WorkOrder's data into a properly formatted .xlsx file."""
        twois_template = load_workbook(TEMPLATE_TWOIS)
        twois_sheet = twois_template.active

        # CONSTANT VALUES    
        twois_sheet['B4'] = 'RS'
        twois_sheet['B5'] = 'BC'
        twois_sheet['C7'] = 'S'
        twois_sheet['B10'] = 'Brian Cobb' #MAKE_CONFIGURABLE
        twois_sheet['G10'] = 'N/A'

        # TWOIS FORM OPTIONS
        twois_sheet['B3'] = this.planned_date.strftime("%#m/%#d/%Y")
        twois_sheet['B3'].alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')
        twois_sheet['A7'] = "" if this.number.startswith("Pending") else this.number[:6]
        twois_sheet['B7'] = this.site_string
        twois_sheet['D7'] = this.title
        twois_sheet['I7'] = this._generate_wotype_string()
        twois_sheet['D8'] = this.location_string
        twois_sheet['A10'] = this.priority_index + 1
        twois_sheet['I10'] = "PAC - Required" if this.requires_pac else "PAC - No"

        # ADVANCED SETTINGS
        twois_sheet['A12'] = "Yes" if this.requires_ncr else "No" 
        twois_sheet['B12'] = this.ncr_string if this.requires_ncr else "N/A"
        twois_sheet['E12'] = "Yes" if this.requires_task_lead else "No" 
        twois_sheet['F12'] = "Yes" if this.tech_witness_point else "No" 
        twois_sheet['G12'] = "Yes" if this.requires_peer_review else "No" 
        twois_sheet['H12'] = "Yes" if this.peer_review_attached else "No" 
        twois_sheet['I12'] = "Yes" if this.requires_ehs else "No" 
        twois_sheet['J12'] = "Yes" if this.qamip else "No" 
        twois_sheet['K12'] = "Yes" if this.requires_qa_review else "No" 

        # TASK LIST
        taskNo: int = 0
        xlrow: int = 15

        for task in this.tasks:
            twois_sheet[f'A{str(xlrow)}'] = task.number
            twois_sheet[f'B{str(xlrow)}'] = task.summary
            twois_sheet[f'G{str(xlrow)}'] = task.reference
            taskNo += 10
            xlrow += 1
        
        twois_template.save(this.filepath)
        twois_template.close()

    
    def save_workorder_as_twois_file(this) -> None:
        """Method which serializes the WorkOrder object as a .twois file."""
        with open(F"{IN_PROGRESS_DIR}\\{this.number}.twois", "wb") as outfile:
            pickle.dump(this, outfile)


    def is_pending(this) -> bool:
        """Method to determine whether or not a work order is pending.\n
        Determines this simply by examining the Work Order number stored in the object."""
        return this.number.startswith("Pending")
        #BCOBB: Maybe I want to do do a little more footwork here, like do a regex on the WO number for the six digits, and if it doesn't start with six digits then return False, maybe?
    

    def reschedule(this, newDate: date) -> None:
        """Method to change the working date of a work order."""
        this.planned_date = newDate
        wb: Workbook = load_workbook(this.filepath)
        ws = wb.active
        ws['B3'] = this.planned_date.strftime("%#m/%#d/%Y")
        wb.save(this.filepath)
        wb.close()
        this.save_workorder_as_twois_file()

    
    def set_file_path(this, newpath: str) -> None:
        this.filepath = newpath
        this.save_workorder_as_twois_file()

    def get_date_string(this) -> str:
        return this.planned_date.strftime('%#m/%#d/%Y')


   

    ############ INTERNAL METHODS ###########################################################################
    def _determine_workorder_type_index(this, cell: str) -> int:
        """Method which determines the Work Order \"type\" combo box index based on the associated cell in the excel worksheet.\n
        This method requires the value of the cell in location I7 on the spreadsheet."""
        resp: int = -1
        for k, v in G_WO_TYPE_DICT_ABBR.items():
            if cell == k:
                resp = v
        return resp
    

    def _determine_location_index(this, base: str, room: str) -> int:
        """Method which returns the Work Order \"location\" combo box index based on the associated cells in the excel worksheet.\n
        This method requires the value of the cell in locations B7("base") and D8("room") on the spreadsheet."""
        resp: int = 0
        #Note: I'm passing the base in because I should but I don't actually use this information yet. Maybe someday it will be used.
        bldg1768reg = re.compile(r"1768")
        bldg = bldg1768reg.search(room)
        if bldg != None:                        # this means we're in building 1768
            roomregex = re.compile(r'21')
            roomno = roomregex.search(room)
            if roomno != None:                  # this means we're in room 21
                resp = 0
            else:
                roomregex = re.compile(r'2')
                roomno = roomregex.search(room)
                if roomno != None:              # this means we're in room 2
                    resp = 2
                else:                           # We're probably in room 6/7/8
                    resp = 1
        return resp
    

    def _get_task_list_from_excel_sheet(this, sheet) -> list[TaskItem]:
        """Method which builds a list of TaskItem objects from the excel spreadsheet by taking the values from rows 15 through 26 and storing their strings in an object.\n
        This method takes in an openpyxl \"worksheet\" object."""
        resp: list[TaskItem] = []
        for i in range(15, 30):
            taskNo: str = "A" + str(i)
            summary: str = "B" + str(i)
            reference: str = "G" + str(i)
            if (sheet[summary].value != None):
                resp.append(TaskItem(sheet[taskNo].value, sheet[summary].value, sheet[reference].value, i))
        return resp
    
    
    def _generate_wotype_string(this) -> str:
        """Method that gets the work order type string from a global constant dictionary based on the stored index."""
        types: list[str] = list(G_WO_TYPE_DICT_ABBR.keys())
        return types[this.workorder_type_index]