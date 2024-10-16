from datetime import date, datetime
from openpyxl import Workbook, load_workbook            #type:ignore
from openpyxl.worksheet.worksheet import Worksheet      #type:ignore
from tkinter import *
from tkinter import filedialog, messagebox
from typing import Callable

import os, pickle, shutil

from lib.forms.dateselector import DateSelectForm
from lib.forms.edit_workorder import EditWorkOrderForm
from lib.global_resources import *
from lib.forms.new_workorder import NewWorkOrderForm
from lib.forms.tkform_wrappers import Tk_Adapter
from lib.forms.generic_workorders import GenericWorkordersForm
from lib.workorder import WorkOrder
from lib.forms.close_workorder import CloseWorkOrder
from lib.outlook_utils import generate_calendar_event_for_approved_work

class MainWindow(Tk_Adapter):
    def __init__(this) -> None:
        Tk_Adapter.__init__(this)

        this.title('B.Cobb\'s TWOIS Manager')
        this.config(background = "white")
        this.active_workorder: WorkOrder
        this.datevar: StringVar = StringVar(None, value="Select")
        this.__date_set: bool = False

        #Create and define the menu bar
        this.menubar: Menu = Menu(this)
        this.menu_generate_twois: Menu = Menu(this.menubar, tearoff=0)
        this.menu_generate_twois.add_command(label = "Generic TWOIS", command = this.generate_generic_workorder)
        this.menu_generate_twois.add_command(label = "Custom TWOIS", command = this.generate_custom_workorder)
        this.menu_generate_twois.add_command(label = "Load Approved TWOIS", command = this.get_approved_workorder)
        this.menubar.add_cascade(label = "Generate TWOIS", menu=this.menu_generate_twois)
        this.menubar.add_command(label = "Configure", command = this.configure_app)

        #Create and define the status bar
        this.footerframe: Frame = Frame(this)
        this.footerframe.pack(side='bottom', fill='x')
        this.statusbar: Label = Label(this.footerframe, text="Ready")
        this.statusbar.bind("<Button-1>", lambda e:this.statusbar.configure(text = "Ready"))
        this.statusbar.pack(side='left')

        #Create and define the TWOIS list
        this.workorder_top_frame: Frame = Frame(this, background = 'white')
        this.workorder_top_frame.pack(side='top')
        this.workorder_frame: Frame = Frame(this.workorder_top_frame, background = "white")
        this.workorder_frame.pack(side='left', padx=20)
        this.build_workorder_table()

        #Create and define the details box
        this.full_details_frame: Frame = Frame(this, bg='white')

        this.button_frame: Frame = Frame(this.full_details_frame, bg='white')
        this.approve_btn: Button = Button(this.button_frame, text="Approve", width=12, command=this.approve_existing_workorder)
        this.reschedule_btn: Button = Button(this.button_frame, text="Reschedule", width=12, command=this.reschedule_workorder)
        this.edit_btn: Button = Button(this.button_frame, text="Edit", width=12, command=this.edit_workorder)
        this.complete_btn: Button = Button(this.button_frame, text="Complete", width=12, command=this.complete_and_send_workorder)
        this.delete_btn: Button = Button(this.button_frame, text="Delete", width=12, command=this.delete_workorder)
        this.hide_btn: Button = Button(this.button_frame, text="Hide", width=12, command=this.hide_workorder_details)

        this.details_frame: Frame = Frame(this.full_details_frame, bg='white')
        this.details_lbl: Label = Label(this.details_frame, width=105, justify='left', bg='white', anchor='nw')
        this.tasklist_frame: Frame = Frame(this.details_frame, bg='white')
        this.tasknums_lbl: Label = Label(this.tasklist_frame, width=15, justify='left', bg='white', anchor='sw', pady=20)
        this.tasklist_lbl: Label = Label(this.tasklist_frame, width=45, justify='left', bg='white', anchor='sw', pady=20)
        this.taskrefs_lbl: Label = Label(this.tasklist_frame, width=45, justify='left', bg='white', anchor='sw', pady=20)

        #Put the window together
        WINDOW_WIDTH: int = 890
        WINDOW_HEIGHT: int = NUMROWS * 80 + 500
        DIMSTR: str = str(WINDOW_WIDTH) + "x" + str(WINDOW_HEIGHT)
        this.geometry(DIMSTR)
        this.resizable(False, True)

        this.config(menu = this.menubar)



    def clear_workorder_table(this) -> None:
        for child in this.workorder_frame.winfo_children():
            child.destroy()

        for j in range(3):
            next_entry: Entry = Entry(this.workorder_frame)
            next_entry.grid(row = 0, column = j)
            val: str = ""
            match j:
                case 0:
                    val = "TWOIS #"
                    next_entry.config(width=14, font=('Arial', FONTSIZE, 'bold'))
                case 1:
                    val = "Work Order Description"
                    next_entry.config(width=86, font=('Arial', FONTSIZE, 'bold'))
                case 2:
                    val = "Planned Due Date"
                    next_entry.config(width=20, font=('Arial', FONTSIZE, 'bold'))
            next_entry.insert(END, val)


    def load_dot_twois_files(this) -> list[WorkOrder]:
        workorders: list[WorkOrder] = []
        for name in os.listdir(IN_PROGRESS_DIR):
            if name.endswith(".twois"):
                file: str = IN_PROGRESS_DIR + "\\" + name
                with open(file, "rb") as infile:
                    new_workorder: WorkOrder = pickle.load(infile)
                    workorders.append(new_workorder)
        return workorders




    def build_workorder_table(this, *args) -> None:
        this.clear_workorder_table()
        workOrders: list[WorkOrder] = this.load_dot_twois_files()

        # Generate the table with clickable events
        for i in range(len(workOrders)):
            NUMROWS: int = i + 2
            for j in range(3):
                next_entry: Entry = Entry(this.workorder_frame, cursor='tcross')
                next_entry.grid(row = i + 1, column = j)
                val: str = ""
                match j:
                    case 0:
                        val = workOrders[i].number
                        next_entry.config(width=14, font=('Arial', FONTSIZE))
                    case 1:
                        val = workOrders[i].title
                        next_entry.config(width=86, font=('Arial', FONTSIZE))
                    case 2:
                        val = str(workOrders[i].get_date_string())
                        next_entry.config(width=20, font=('Arial', FONTSIZE))

                next_entry.bind("<Button-1>", this.post_workorder_details_with_lambda(this.statusbar, workOrders[i]))
                next_entry.insert(END, val)


    #BCOBB: Could go into an excel_utils.py file??
    def file_is_an_approved_workorder(this, file: str) -> bool:
        if not file.endswith(".xlsx"):
            this.statusbar.configure(text="Selected file is not a .xlsx file!")
            return False
        
        workbook: Workbook = load_workbook(filename=file)
        sheet: Worksheet = workbook.active
        if (sheet["D1"].value != 'Technician Work Order Information Sheet') or sheet['A13'].value != 'Work Order Plans':
            this.statusbar.configure(text="Selected file is not a TWOIS!")
            workbook.close()
            return False

        print(f"{str(sheet['B3'].value)}   {str(sheet['B4'].value)}   {str(sheet['B5'].value)}")
        if (sheet['B3'].value == None) or (sheet['B4'].value == None) or (sheet['B5'].value == None):
            this.statusbar.configure(text="Selected TWOIS hasn't been approved by Production Control!")
            workbook.close()
            return False
        
        if (not len(str(sheet['A7'].value)) == 6 or not str(sheet['A7'].value).isnumeric()):
            this.statusbar.configure(text="Selected TWOIS has no work order number!")
            workbook.close()
            return False
        
        workbook.close()
        return True
    
    
    #BCOBB: Could go into an excel_utils.py file??
    def file_does_not_match_workorder(this, file: str, workorder: WorkOrder):
        workbook: Workbook = load_workbook(filename=file)
        sheet: Worksheet = workbook.active
        if sheet['D7'].value != workorder.title:
            this.statusbar.configure(text="Selected file is not a TWOIS!")
            workbook.close()
            return False
        return True


    def post_workorder_details_with_lambda(this, statusbar: Label, workOrder: WorkOrder) -> Callable:
        return lambda ev:this.post_workorder_details(statusbar, workOrder)


    def hide_workorder_details(this, *args) -> None:
        this.approve_btn.pack_forget()
        this.reschedule_btn.pack_forget()
        this.edit_btn.pack_forget()
        this.complete_btn.pack_forget()
        this.delete_btn.pack_forget()
        this.hide_btn.pack_forget()
        this.button_frame.pack_forget()
        this.tasklist_frame.pack_forget()
        this.tasknums_lbl.pack_forget()
        this.tasklist_lbl.pack_forget()
        this.taskrefs_lbl.pack_forget()
        this.details_lbl.pack_forget()
        this.details_frame.pack_forget()
        this.full_details_frame.pack_forget()
        this.statusbar.configure(text="Ready")


    def show_workorder_details(this) -> None:        
        this.approve_btn.pack(side='top', pady=5)
        this.reschedule_btn.pack(side='top', pady=5)
        this.edit_btn.pack(side='top', pady=5)
        this.complete_btn.pack(side='top', pady=5)
        this.delete_btn.pack(side='top', pady=5)
        this.hide_btn.pack(side='top', pady=5)
        this.button_frame.pack(side='right', pady=5, anchor='ne')
        this.details_lbl.pack(side='top')
        this.tasklist_frame.pack(side='bottom')
        this.tasknums_lbl.pack(side='left')
        this.tasklist_lbl.pack(side='left')
        this.taskrefs_lbl.pack(side='left')
        this.details_frame.pack(side='left')
        this.full_details_frame.pack(side='top', pady=15)


    def post_workorder_details(this, statusbar: Label, workorder: WorkOrder) -> None:
        this.hide_workorder_details()
        this.active_workorder = workorder

        statusbar.configure(text = f'{workorder.number} selected')
        if workorder.is_pending():
            this.approve_btn.configure(state='normal')
            this.complete_btn.configure(state='disabled')
        else:
            this.approve_btn.configure(state='disabled')
            this.complete_btn.configure(state='normal')

        task_numbers: str = f"Task No.:"
        task_list: str = f"Task Summary:"
        task_refs: str = f"Reference Docs:"
        for task in workorder.tasks:
            task_numbers += f"\nTask {str(task.number)}"
            task_list += f"\n{task.summary}"
            task_refs += f"\n{task.reference}"

        this.tasknums_lbl.configure(text=task_numbers)
        this.tasklist_lbl.configure(text=task_list)
        this.taskrefs_lbl.configure(text=task_refs)
        details_string: str = f"{workorder.number}\t\t{workorder.title}\n\nDue {str(workorder.planned_date)}\n\n{workorder.description}"
        this.details_lbl.configure(text=details_string)

        this.show_workorder_details()

    

    def generate_custom_workorder(this, *args) -> None:
        new_workorder: NewWorkOrderForm = NewWorkOrderForm(this)
        new_workorder.bind("<Destroy>", this.build_workorder_table)



    def get_approved_workorder(this, *args) -> None:
        filepath: str = this.get_approved_workorder_from_user()

        if this.file_is_an_approved_workorder(filepath):
            #BCOBB: Add function "workOrder.getDescription"     # <-- this method should get a task description from the user.
            workorder: WorkOrder = WorkOrder(filepath)
            fsplit: list[str] = filepath.split('/')
            filename = fsplit[len(fsplit) - 1]
            newpath: str = f"{IN_PROGRESS_DIR}\\{filename}"
            shutil.copyfile(filepath, newpath)
            workorder.set_file_path(newpath)
            workorder.save_workorder_as_twois_file()
            this.statusbar.configure(text="Copied file: \"" + filepath + "\" to the working TWOIS directory")
        
        this.build_workorder_table()



    def complete_and_send_workorder(this, *args) -> None:
        completeforms = CloseWorkOrder(this, this.active_workorder)
        completeforms.bind("<Destroy>", this.reset_workorder_display)


    def approve_existing_workorder(this, *args) -> None:
        #BCOBB: DEFINE THIS FUNCTION
        filepath = this.get_approved_workorder_from_user()

        if not this.file_is_an_approved_workorder(filepath):
            return
        
        if not this.file_does_not_match_workorder(filepath, this.active_workorder):
            return
        
        workorder: WorkOrder = WorkOrder(filepath)
        fsplit: list[str] = filepath.split('/')
        filename = fsplit[len(fsplit) - 1]
        newpath: str = f"{IN_PROGRESS_DIR}\\{filename}"
        shutil.copyfile(filepath, newpath)
        workorder.set_file_path(newpath)
        workorder.save_workorder_as_twois_file()
        this.statusbar.configure(text="Copied file: \"" + filepath + "\" to the working TWOIS directory")
        
        os.remove(this.active_workorder.filepath)
        os.remove(f"{IN_PROGRESS_DIR}\\{this.active_workorder.number}.twois")
        this.active_workorder = workorder

        generate_calendar_event_for_approved_work(this.active_workorder, newpath)
        this.reset_workorder_display()
        print("Need to define updating a TWOIS to approved status!")


    def reschedule_workorder(this, *args) -> None:
        dsform = DateSelectForm(this.datevar)
        dsform.bind("<Destroy>", this._reschedule_workorder)

    def _reschedule_workorder(this, event: Event) -> None:        
        #BCOBB: Learned lesson--set all ".bind" callables to execute only once, using the if statement in this method
        if event.widget == event.widget.winfo_toplevel():
            dsplit: list[str] = this.datevar.get().split("/")
            dateresp: date = date(int(dsplit[2]), int(dsplit[0]), int(dsplit[1]))
            this.active_workorder.reschedule(dateresp)
            this.reset_workorder_display()


    def edit_workorder(this, *args) -> None:
        form: EditWorkOrderForm
        if this.active_workorder != None:
            form = EditWorkOrderForm(this, this.active_workorder)
            form.bind("<Destroy>", this.reset_workorder_display)
        else:
            this.statusbar.configure(text="No valid Work Order is selected!")




    def delete_workorder(this, *args) -> None:
        user_definitely_for_sure_wants_to_delete: bool = messagebox.askyesno(f"Delete Work Order {this.active_workorder.number}", f"Are you sure that you want to delete TWOIS# {this.active_workorder.number}?\nNote that, once deleted, it will be gone forever.")
        if user_definitely_for_sure_wants_to_delete:
            if os.path.exists(this.active_workorder.filepath):
                os.remove(this.active_workorder.filepath)
            if os.path.exists(f"{IN_PROGRESS_DIR}\\{this.active_workorder.number}.twois"):
                os.remove(f"{IN_PROGRESS_DIR}\\{this.active_workorder.number}.twois")
            this.reset_workorder_display()
        else:
            this.statusbar.configure(text=f"TWOIS# {this.active_workorder.number} is safe for another day!")


    def configure_app(this, *args) -> None:
        #BCOBB
        print("Need to define a configuration file and implement it!")


    def generate_generic_workorder(this, *args) -> None:
        GenericWorkordersForm(this)
        #BCOBB
        print("Need to define generic TWOIS templates!")


    def reset_workorder_display(this, *args) -> None:
        this.hide_workorder_details()
        this.build_workorder_table()



    def get_approved_workorder_from_user(this) -> str:
        filepath: str = filedialog.askopenfilename(initialdir=USERHOME, title = "Select a TWOIS",
                                        filetypes = ( ("Excel files (*.xlsx)", "*.xlsx*"), ("All files", "*.*") ))
        return filepath


