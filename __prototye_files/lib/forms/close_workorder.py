from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox
from datetime import date
from lib.workorder import WorkOrder
from lib.forms.tkform_wrappers import Toplevel_Adapter
from lib.custom_tkwidgets import DateEntry
from lib.outlook_utils import generate_work_order_completion_email
from lib.global_resources import *
from lib.forms.dateselector import DateSelectForm
from openpyxl import load_workbook, Workbook        #type:ignore
from openpyxl.worksheet.worksheet import Worksheet  #type:ignore

import os

class CloseWorkOrder(Toplevel_Adapter):
    def __init__(this, master: Toplevel | Tk, workorder: WorkOrder) -> None:
        Toplevel_Adapter.__init__(this, master)
        this.title(f"Complete TWOIS# {workorder.number}")
        this.geometry("700x200")
        this.workorder = workorder

        task_table_frame: Frame = Frame(this)

        head_number: Entry = Entry(task_table_frame)
        head_number.insert(0, "Task No.")
        head_number.config(width=10, font=('Arial', FONTSIZE, 'bold'), state='readonly')
        head_number.grid(row=0, column=0)

        head_summary: Entry = Entry(task_table_frame)
        head_summary.insert(0, "Task Summary")
        head_summary.config(width=40, font=('Arial', FONTSIZE, 'bold'), state='readonly')
        head_summary.grid(row=0, column=1)

        head_admins: Entry = Entry(task_table_frame)
        head_admins.insert(0, "Technician")
        head_admins.config(width=20, font=('Arial', FONTSIZE, 'bold'), state='readonly')
        head_admins.grid(row=0, column=2)

        head_techs: Entry = Entry(task_table_frame)
        head_techs.insert(0, "# Techs")
        head_techs.config(width=10, font=('Arial', FONTSIZE, 'bold'), state='readonly')
        head_techs.grid(row=0, column=3)
        
        head_hours: Entry = Entry(task_table_frame)
        head_hours.insert(0, "# Hours")
        head_hours.config(width=10, font=('Arial', FONTSIZE, 'bold'), state='readonly')
        head_hours.grid(row=0, column=4)

        this.task_tech_boxes: list[Spinbox] = []
        this.task_hours_boxes: list[Spinbox] = []
        this.admin_boxes: list[Combobox] = []

        i: int = 1
        for task in this.workorder.tasks:
            task_number: Entry = Entry(task_table_frame)
            task_number.insert(0, str(task.number))
            task_number.config(width=10, font=('Arial', FONTSIZE), state='readonly')
            task_number.grid(row=i, column=0)

            task_summary: Entry = Entry(task_table_frame)
            task_summary.insert(0, str(task.summary))
            task_summary.config(width=40, font=('Arial', FONTSIZE), state='readonly')
            task_summary.grid(row=i, column=1)
            
            admin_select: Combobox = Combobox(task_table_frame, width=18, height=1)
            for v in g_admin_dict.values():
                admin_select['values'] = (*admin_select['values'], v)
            admin_select.current(0)
            admin_select.config(font=('Arial', FONTSIZE), justify='center', state='readonly')
            admin_select.grid(row=i, column=2)
            this.admin_boxes.append(admin_select)

            task_techs: Spinbox = Spinbox(task_table_frame, from_=1, to=10, increment=1)
            task_techs.config(width=8, font=('Arial', FONTSIZE), justify='center', state='readonly', readonlybackground='white')
            task_techs.grid(row=i, column=3)
            this.task_tech_boxes.append(task_techs)
            
            task_hours: Spinbox = Spinbox(task_table_frame, from_=0.1, to=100, increment=0.1)
            task_hours.config(width=8, font=('Arial', FONTSIZE), justify='center', state='readonly', readonlybackground='white')
            task_hours.grid(row=i, column=4)
            this.task_hours_boxes.append(task_hours)
            i += 1
        
        task_table_frame.pack(side='top', pady=20)

        date_frame: Frame = Frame(this)
        date_label: Label = Label(date_frame, text='Date of Completion:', anchor='w')
        this.date_entry: DateEntry = DateEntry(date_frame)

        date_label.pack(side='left', padx=20)
        this.date_entry.pack(side='right', padx=20)
        date_frame.pack(side='top', pady=10)

        buttons_frame: Frame = Frame(this)
        button_submit: Button = Button(buttons_frame, text="Submit", width=15, justify='center', command=this._submit)
        button_cancel: Button = Button(buttons_frame, text="Cancel", width=15, justify='center', command=this._cancel)

        button_submit.pack(side='left', padx=20)
        button_cancel.pack(side='right', padx=20)
        buttons_frame.pack(side='top', pady=10)

        this.grab_set()

    def _cancel(this) -> None:
        this.destroy()


    def _submit(this) -> None:
        if not this._all_inpus_are_valid():
            return
        
        #BCOBB: This could be a "WorkOrder.complete_workorder" method, owned by the WorkOrder class?
        wb: Workbook = load_workbook(this.workorder.filepath)
        ws: Worksheet = wb.active
        
        for i in range(len(this.workorder.tasks)):
            task = this.workorder.tasks[i]
            ws[f'A{task.actuals_row}'] = task.number
            ws[f'C{task.actuals_row}'] = this.get_date_string()
            for k, v in g_admin_dict.items():
                if this.admin_boxes[i].get() == v:
                    ws[f'E{task.actuals_row}'] = f"{v} / {k}"
            ws[f'H{task.actuals_row}'] = this.task_tech_boxes[i].get()
            ws[f'J{task.actuals_row}'] = this.task_hours_boxes[i].get()

        d: date = this.get_date()
        date_dir: str = f"{COMPLETE_DIR}\\{d.strftime('%Y%m%d')}"
        fname: str = f"{this.workorder.number} - {this.workorder.title}.xlsx"
        if not os.path.exists(date_dir):
            os.makedirs(date_dir)
        
        wb.save(f"{date_dir}\\{fname}")
        wb.close()
        
        generate_work_order_completion_email(this.workorder, f"{date_dir}\\{fname}")

        description_file = open(f"{date_dir}\\{this.workorder.number}.description.txt", 'w')
        description_file.write(this.workorder.description)
        description_file.close()

        os.remove(f"{IN_PROGRESS_DIR}\\{this.workorder.number}.twois")
        os.remove(this.workorder.filepath)
        this.destroy()
    
    def _all_inpus_are_valid(this) -> bool:
        """Method used for determining that the form is properly filled out before proceeding with work order completion."""
        completion_date: str = this.date_entry.get()
        if len(completion_date) > 10 or completion_date.startswith("Select") or len(completion_date.split("/")) != 3:
            messagebox.showerror("Date Select Error", "Error: A completion date must be selected")
            return False
        return True






