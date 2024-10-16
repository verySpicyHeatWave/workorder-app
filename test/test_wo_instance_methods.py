"""This test file is meant to ensure that a workorder object being generated from an Excel
spreadsheet will correctly parse and populate its fields based on the contents of that
spreadsheet.
"""

#pylint: skip-file

from datetime import date, timedelta
import os
import random
import sys
import unittest

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from appfiles.library.special import Special
from appfiles.utils.utils import date_to_string

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from appfiles.library.excelfilestatus import ExcelFileStatus
from appfiles.library.logcomment import LogComment
from appfiles.library.site import Site
from appfiles.library.taskitem import TaskItem
from appfiles.library.workorder import WorkOrder
import appfiles.utils.appglobals as g #type: ignore
from appfiles.utils.utils_for_testing import clear_dir, get_date_dir_name, insert_approval_markings_on_wo, insert_approval_stamps
from appfiles.utils.workorder import approved_file_status, extract_pending_twois_number, get_room_from_xlsx_cell, is_a_valid_wo_number


class WorkOrderSaveTests(unittest.TestCase):
    """Defines the tests for use with all of the save methods, including save(),
    save_as_excel_file(), save_as_twois_file(), and save_as_wo_template().
    """
    def test_save__files_are_created(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder()
        test_wo.save()
        self.assertTrue(os.path.isfile(test_wo.get_twois_filepath()))
        self.assertTrue(os.path.isfile(test_wo.get_excel_filepath()))

    def test_save__files_are_pending1(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder()
        test_wo.save()
        lsdir: list[str] = os.listdir(g.IN_PROGRESS_DIR)
        self.assertEqual(len(lsdir), 3) #One for the test file, two for the WO files
        self.assertTrue("Pending-001.twois" in lsdir)
        self.assertTrue(f"Pending-001 - {test_wo.title}.xlsx" in lsdir)

    def test_save__files_are_pending2(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)        
        for i in range(8):
            wo = WorkOrder()
            wo.save()
        rmfile :str = os.path.join(g.IN_PROGRESS_DIR, "Pending-005.twois")
        os.remove(rmfile)
        test_wo: WorkOrder = WorkOrder()
        test_wo.save()
        lsdir: list[str] = os.listdir(g.IN_PROGRESS_DIR)
        self.assertEqual(len(lsdir), 17) #One for the test file, two each for the WO files (8 total)
        self.assertTrue("Pending-005.twois" in lsdir)
        self.assertTrue(f"Pending-005 - {test_wo.title}.xlsx" in lsdir)

    def test_save__files_have_validated_wo_number(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder(wo_number="123456", site=Site.FG, special=Special.I)
        test_wo.save()
        wonum: str = "123456FGI"
        lsdir: list[str] = os.listdir(g.IN_PROGRESS_DIR)
        self.assertTrue(f"{wonum}.twois" in lsdir)
        self.assertTrue(f"{wonum} - {test_wo.title}.xlsx" in lsdir)

    def test_save__properly_uses_existing_file(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder(wo_number="123456", site=Site.FG, special=Special.I)
        test_wo.save()
        wofile: str = test_wo.get_excel_filepath()
        wb: Workbook = load_workbook(wofile)
        ws: Worksheet = wb.active
        ws['C3'].value = "pickle"
        wb.save(wofile)
        wb.close()
        test_wo.edit(description="Hello you fool!")
        wb2 = load_workbook(wofile)
        ws2 = wb.active
        value = str(ws2['C3'].value)
        wb2.save(wofile)
        wb2.close()
        self.assertEqual(value, "pickle")

    def test_save__wo_cell_is_blank_if_pending(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder(site=Site.FG, special=Special.I)
        test_wo.save()
        wofile: str = test_wo.get_excel_filepath()
        wb: Workbook = load_workbook(wofile)
        ws: Worksheet = wb.active
        value = str(ws['A7'].value)
        wb.save(wofile)
        wb.close()
        self.assertEqual(value, "None")

    def test_save__wo_title_is_appropriate(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlemonkey: str = "Ook Ook"
        test_wo: WorkOrder = WorkOrder(title=titlemonkey)
        test_wo.save()
        lsdir: list[str] = os.listdir(g.IN_PROGRESS_DIR)
        self.assertTrue(f"Pending-001 - {titlemonkey}.xlsx" in lsdir)
        self.assertEqual(titlemonkey, test_wo.title)

    def test_save_temp__saves_successfully(self) -> None:
        clear_dir(g.TEMPLATE_DIR)
        titlemonkey: str = "Ook Ook"
        test_wo: WorkOrder = WorkOrder(title=titlemonkey)
        test_wo.save_as_wo_template('Ook Ook.twois')
        file = os.path.join(g.TEMPLATE_DIR, 'Ook Ook-1.twois')
        self.assertTrue(os.path.isfile(file))

    def test_save_temp__saves_with_multiples(self) -> None:
        clear_dir(g.TEMPLATE_DIR)
        titlemonkey: str = "Ook Ook"
        test_wo: WorkOrder = WorkOrder(title=titlemonkey)
        for i in range(4):
            test_wo.save_as_wo_template('Ook Ook')
        file = os.path.join(g.TEMPLATE_DIR, 'Ook Ook-4.twois')
        self.assertTrue(os.path.isfile(file))

    def test_save_temp__overwrites_existing_file(self) -> None:
        titlemonkey: str = "Ook Ook"
        test_wo: WorkOrder = WorkOrder(title=titlemonkey)
        file = os.path.join(g.TEMPLATE_DIR, 'Ook Ook-2.twois')
        test_wo.save_as_wo_template(file)
        modtime: float = 0
        oldestfile: str = ""
        for f in os.listdir(g.TEMPLATE_DIR):
            ff = os.path.join(g.TEMPLATE_DIR, f)
            newmodtime: float = os.path.getmtime(ff)
            if newmodtime > modtime:
                modtime = newmodtime
                oldestfile = ff
        self.assertEqual(oldestfile, file)

    def test_save_temp__creates_new_file_with_full_path(self) -> None:
        titlemonkey: str = "Ook Ook"
        test_wo: WorkOrder = WorkOrder(title=titlemonkey)
        file = os.path.join(g.TEMPLATE_DIR, 'Ook Ook-275.twois')
        test_wo.save_as_wo_template(file)
        self.assertTrue(os.path.isfile(file))



class WorkOrderEditTests(unittest.TestCase):
    """Defines the tests for use with the edit() method."""
    def test_edit__field_updates_correctly1(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        firsttitle: str = "Hello monkey"
        secondtitle: str = "Eat banana"
        test_wo: WorkOrder = WorkOrder(title=firsttitle)
        self.assertEqual(test_wo.get_excel_filepath(), f"{g.IN_PROGRESS_DIR}\\Pending-001 - {firsttitle}.xlsx")
        test_wo.edit(title=secondtitle)
        self.assertEqual(test_wo.get_excel_filepath(), f"{g.IN_PROGRESS_DIR}\\Pending-001 - {secondtitle}.xlsx")

    def test_edit__field_updates_correctly2(self) -> None:
        _num: str = "123456"
        _site: Site = Site.FG
        _spec: Special = Special.S
        result: str = f"{_num}{_site.name}{_spec.name}"
        test_wo: WorkOrder = WorkOrder(wo_number=_num, site=_site, special=_spec)
        self.assertEqual(test_wo.get_full_workorder_number(), result)
        test_wo.edit(site=Site.VB)
        result = f"{_num}{Site.VB.name}{_spec.name}"
        self.assertEqual(test_wo.get_full_workorder_number(), result)

    def test_edit__unspecd_fields_donot_update(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        firsttitle: str = "Hello monkey"
        secondtitle: str = "Eat banana"
        test_wo: WorkOrder = WorkOrder(title=firsttitle, room=255)
        self.assertEqual(test_wo.room, 255)
        test_wo.edit(title=secondtitle)
        self.assertEqual(test_wo.room, 255)

    def test_edit__propagate_into_excel(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        room1 = 21
        room2 = 255
        test_wo: WorkOrder = WorkOrder(room=room1)
        test_wo.save()
        wb: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws: Worksheet = wb.active
        value = get_room_from_xlsx_cell(ws)
        wb.close()
        self.assertEqual(value, room1)
        test_wo.edit(room=room2)
        wb2 = load_workbook(test_wo.get_excel_filepath())
        ws2 = wb2.active
        value = get_room_from_xlsx_cell(ws2)
        wb2.close()
        self.assertEqual(value, room2)

    def test_edit__filenames_change_accordingly1(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        firsttitle: str = "Hello monkey"
        secondtitle: str = "Eat banana"
        test_wo: WorkOrder = WorkOrder(title=firsttitle, room=255, site=Site.VB, special=Special.S)
        test_wo.save()
        lsdir: list[str] = os.listdir(g.IN_PROGRESS_DIR)
        self.assertTrue(f"Pending-001.twois" in lsdir)
        self.assertTrue(f"Pending-001 - {firsttitle}.xlsx" in lsdir)
        test_wo.edit(wo_number="123456", title=secondtitle)
        lsdir = os.listdir(g.IN_PROGRESS_DIR)
        self.assertTrue(f"123456VBS.twois" in lsdir)
        self.assertTrue(f"123456VBS - {secondtitle}.xlsx" in lsdir)
        self.assertFalse(f"Pending-001.twois" in lsdir)
        self.assertFalse(f"Pending-001 - {firsttitle}.xlsx" in lsdir)
        


class WorkOrderDeleteTests(unittest.TestCase):
    def test_delete__make_sure_file_is_deleted(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder()
        test_wo.save()
        self.assertTrue(os.path.exists(test_wo.get_twois_filepath()))
        self.assertTrue(os.path.exists(test_wo.get_excel_filepath()))
        test_wo.delete()
        self.assertFalse(os.path.exists(test_wo.get_twois_filepath()))
        self.assertFalse(os.path.exists(test_wo.get_excel_filepath()))


class WorkOrderApproveTests(unittest.TestCase):
    """Defines the tests for use with the approve() method."""

    def test_approve__workorder_unchanged_with_bad_file1(self) -> None:
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, title="It is a brave new world out there")
        resp: ExcelFileStatus = test_wo.approve('invalid_file.txt')
        self.assertEqual(resp, ExcelFileStatus.NOT_XLSX)
        self.assertEqual(test_wo.title, "It is a brave new world out there")
        self.assertTrue(test_wo.wo_number.startswith("Pending"))
        pass

    
    def test_approve__workorder_unchanged_with_bad_file2(self) -> None:
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, title="It is a brave new world out there")
        resp: ExcelFileStatus = test_wo.approve('invalid_file1.xlsx')
        self.assertEqual(resp, ExcelFileStatus.NOT_TWOIS)
        self.assertEqual(test_wo.title, "It is a brave new world out there")
        self.assertTrue(test_wo.wo_number.startswith("Pending"))
        pass

    
    def test_approve__workorder_unchanged_with_bad_file3(self) -> None:
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, title="It is a brave new world out there")
        resp: ExcelFileStatus = test_wo.approve('invalid_file2.xlsx')
        self.assertEqual(resp, ExcelFileStatus.NOT_APPROVED)
        self.assertEqual(test_wo.title, "It is a brave new world out there")
        self.assertTrue(test_wo.wo_number.startswith("Pending"))
    
    
    def test_approve__workorder_unchanged_with_no_file(self) -> None:
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, title="It is a brave new world out there")
        resp: ExcelFileStatus = test_wo.approve('notafile.xlsx')
        self.assertEqual(resp, ExcelFileStatus.NOT_FOUND)
        self.assertEqual(test_wo.title, "It is a brave new world out there")
        self.assertTrue(test_wo.wo_number.startswith("Pending"))

    
    def test_approve__workorder_unchanged_with_unmatched_file(self) -> None:
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, title="It is a brave new world out there")
        test_wo.save()
        resp: ExcelFileStatus = test_wo.approve('unmatched_file.xlsx')
        self.assertEqual(resp, ExcelFileStatus.FILES_UNMATCHED)
        self.assertEqual(test_wo.title, "It is a brave new world out there")
        self.assertTrue(test_wo.wo_number.startswith("Pending"))
    

    def test_approve__is_successful_with_good_file(self) -> None:
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, title="It is a brave new world out there",
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16)])
        wb: Workbook = load_workbook('valid_file.xlsx')
        ws: Worksheet = wb.active
        ws['B3'].value = date_to_string(date.today())
        wb.save('valid_file.xlsx')
        wb.close()
        resp: ExcelFileStatus = test_wo.approve('valid_file.xlsx')
        self.assertEqual(resp, ExcelFileStatus.IS_VALID)
        self.assertEqual(test_wo.title, "It is a brave new world out there")
    

    def test_approve__description_passes_over(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder(site=Site.VB, special=Special.S, description="Jinkies!",
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16)],
                                       title='It is a brave new world out there')
        test_wo.save()
        self.assertTrue(os.path.exists(test_wo.get_twois_filepath()))
        wb: Workbook = load_workbook('valid_file.xlsx')
        ws: Worksheet = wb.active
        ws['B3'].value = date_to_string(date.today())
        wb.save('valid_file.xlsx')
        wb.close()

        resp: ExcelFileStatus = test_wo.approve('valid_file.xlsx')
        self.assertEqual(resp, ExcelFileStatus.IS_VALID)
        self.assertFalse(os.path.exists(test_wo.get_twois_filepath()))

        newfile: str = os.path.join(g.IN_PROGRESS_DIR, '419157VBS.twois')
        self.assertTrue(os.path.exists(newfile))
        new_wo: WorkOrder = WorkOrder.from_twois(newfile)
        self.assertEqual(new_wo.description, "Jinkies!")
    

    def test_approve__fails_if_wo_is_already_approved(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder(wo_number="419157", site=Site.VB, special=Special.S, description="Jinkies!",
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16)],
                                       title='It is a brave new world out there')
        test_wo.save()
        self.assertTrue(os.path.exists(test_wo.get_twois_filepath()))
        wb1: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws1: Worksheet = wb1.active
        insert_approval_stamps(ws1)
        wb1.save(test_wo.get_excel_filepath())
        wb1.close()

        wb2: Workbook = load_workbook('valid_file.xlsx')
        ws2: Worksheet = wb2.active
        ws2['B3'].value = date_to_string(date.today())
        wb2.save('valid_file.xlsx')
        wb2.close()

        resp: ExcelFileStatus = test_wo.approve('valid_file.xlsx')
        self.assertEqual(resp, ExcelFileStatus.WO_ALREADY_APPROVED)
        self.assertTrue(os.path.exists(test_wo.get_twois_filepath()))


class WorkOrderLogCommentTests(unittest.TestCase):
    """Defines the tests for use with the log_comments() method."""
    def test_logcom__comments_append_successfully(self) -> None:
        logComments: list[LogComment] = []
        for i in range(3):
            logComments.append(LogComment(f"Comment number {str(i+1)}", "B Cobb", date.today(), i))
        test_wo: WorkOrder = WorkOrder(comments=logComments)
        self.assertTrue(len(test_wo.comments) == 3)
        for i in range(len(test_wo.comments)):
            self.assertEqual(test_wo.comments[i].text, f"Comment number {str(i+1)}")
            self.assertEqual(test_wo.comments[i].get_row(), i+86)
        self.assertTrue(test_wo.log_comment([LogComment("This is my new comment!", "B Cobb", date.today(), len(test_wo.comments))]))
        self.assertEqual(test_wo.comments[-1].text, f"This is my new comment!")
        self.assertEqual(test_wo.comments[-1].get_row(), 89)


    def test_logcom__comments_append_successfully_with_wrong_row(self) -> None:
        logComments: list[LogComment] = []
        for i in range(3):
            logComments.append(LogComment(f"Comment number {str(i+1)}", "B Cobb", date.today(), i))
        test_wo: WorkOrder = WorkOrder(comments=logComments)
        self.assertTrue(len(test_wo.comments) == 3)
        for i in range(len(test_wo.comments)):
            self.assertEqual(test_wo.comments[i].text, f"Comment number {str(i+1)}")
            self.assertEqual(test_wo.comments[i].get_row(), i+86)
        self.assertTrue(test_wo.log_comment([LogComment("This is my new comment!", "B Cobb", date.today(), 3 + len(test_wo.comments))]))
        self.assertEqual(test_wo.comments[-1].text, f"This is my new comment!")
        self.assertEqual(test_wo.comments[-1].get_row(), 89)


    def test_logcom__comments_do_not_append_if_row_exists(self) -> None:
        logComments: list[LogComment] = []
        for i in range(10):
            logComments.append(LogComment(f"Comment number {str(i+1)}", "B Cobb", date.today(), i))
        test_wo: WorkOrder = WorkOrder(comments=logComments)
        self.assertTrue(len(test_wo.comments) == 10)
        for i in range(len(test_wo.comments)):
            self.assertEqual(test_wo.comments[i].text, f"Comment number {str(i+1)}")
            self.assertEqual(test_wo.comments[i].get_row(), i+86)
        self.assertFalse(test_wo.log_comment([LogComment("This is my new comment!", "B Cobb", date.today(), 6)]))
        for i in range(len(test_wo.comments)):
            self.assertEqual(test_wo.comments[i].text, f"Comment number {str(i+1)}")
            self.assertEqual(test_wo.comments[i].get_row(), i+86)


    def test_logcom__never_exceeds_24(self) -> None:
        logComments: list[LogComment] = []
        for i in range(24):
            logComments.append(LogComment(f"Comment number {str(i+1)}", "B Cobb", date.today(), i))
        test_wo: WorkOrder = WorkOrder(comments=logComments)
        self.assertTrue(len(test_wo.comments) == 24)
        for i in range(len(test_wo.comments)):
            self.assertEqual(test_wo.comments[i].text, f"Comment number {str(i+1)}")
            self.assertEqual(test_wo.comments[i].get_row(), i+86)
        self.assertFalse(test_wo.log_comment([LogComment("This comment shouldn't exist!", "B Cobb", date.today(), len(test_wo.comments))]))
        self.assertTrue(len(test_wo.comments) == 24)
        for i in range(len(test_wo.comments)):
            self.assertEqual(test_wo.comments[i].text, f"Comment number {str(i+1)}")


class WorkOrderCompleteTests(unittest.TestCase):
    """Defines the tests for use with the complete() method."""
    def test_complete__files_are_saved_in_correct_dir(self) -> None:
        clear_dir(g.COMPLETE_DIR)
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder(wo_number="123456",
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        test_wo.save()
        self.assertTrue(os.path.isfile(test_wo.get_excel_filepath()))
        self.assertTrue(os.path.isfile(test_wo.get_twois_filepath()))
        insert_approval_markings_on_wo(test_wo)

        for task in test_wo.task_list:
            task.complete(date.today(), str(g.people[0]), 1, 5 * random.random())
        test_wo.complete(date.today())

        excelname: str = test_wo.get_excel_filepath().split("\\")[-1]

        self.assertFalse(os.path.isfile(test_wo.get_excel_filepath()))
        self.assertFalse(os.path.isfile(test_wo.get_twois_filepath()))
        cmpdir: str = os.path.join(g.COMPLETE_DIR, str(date.today().year))
        self.assertTrue(os.path.isdir(cmpdir))

        cmpdir = os.path.join(cmpdir, get_date_dir_name(date.today()))
        self.assertTrue(os.path.isdir(cmpdir))
        self.assertTrue(os.path.isfile(os.path.join(cmpdir, excelname)))
        self.assertTrue(os.path.isfile(os.path.join(cmpdir,f"{test_wo.get_full_workorder_number()}_description.txt")))




    def test_complete__files_are_saved_in_existing_dir(self) -> None:
        clear_dir(g.COMPLETE_DIR)
        clear_dir(g.IN_PROGRESS_DIR)
        cmpdir: str = os.path.join(g.COMPLETE_DIR, str(date.today().year))
        os.mkdir(cmpdir)
        self.assertTrue(os.path.isdir(cmpdir))
        cmpdir = os.path.join(cmpdir, get_date_dir_name(date.today()))
        os.mkdir(cmpdir)
        self.assertTrue(os.path.isdir(cmpdir))
        test_wo: WorkOrder = WorkOrder(wo_number="123456",
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        test_wo.save()
        insert_approval_markings_on_wo(test_wo)

        for task in test_wo.task_list:
            task.complete(date.today(), str(g.people[0]), 1, 5 * random.random())
        test_wo.complete(date.today())

        excelname: str = test_wo.get_excel_filepath().split("\\")[-1]
        self.assertTrue(os.path.isfile(os.path.join(cmpdir, excelname)))
        self.assertTrue(os.path.isfile(os.path.join(cmpdir,f"{test_wo.get_full_workorder_number()}_description.txt")))

        
    def test_complete__fails_if_workorder_not_approved(self) -> None:
        clear_dir(g.COMPLETE_DIR)
        test_wo: WorkOrder = WorkOrder(task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        test_wo.save()
        self.assertTrue(os.path.isfile(test_wo.get_excel_filepath()))
        self.assertTrue(os.path.isfile(test_wo.get_twois_filepath()))

        for task in test_wo.task_list:
            task.complete(date.today(), str(g.people[0]), 1, 5 * random.random())
        test_wo.complete(date.today())

        excelname: str = test_wo.get_excel_filepath().split("\\")[-1]

        self.assertTrue(os.path.isfile(test_wo.get_excel_filepath()))
        self.assertTrue(os.path.isfile(test_wo.get_twois_filepath()))
        cmpdir: str = os.path.join(g.COMPLETE_DIR, str(date.today().year))
        self.assertFalse(os.path.isdir(cmpdir))
        cmpdir = os.path.join(cmpdir, get_date_dir_name(date.today()))
        self.assertFalse(os.path.isdir(cmpdir))
        self.assertFalse(os.path.isfile(os.path.join(cmpdir, excelname)))
        self.assertFalse(os.path.isfile(os.path.join(cmpdir,f"{test_wo.get_full_workorder_number()}_description.txt")))

        
    def test_complete__fails_if_tasks_are_incomplete(self) -> None:
        clear_dir(g.COMPLETE_DIR)
        test_wo: WorkOrder = WorkOrder(task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        test_wo.save()
        self.assertTrue(os.path.isfile(test_wo.get_excel_filepath()))
        self.assertTrue(os.path.isfile(test_wo.get_twois_filepath()))

        for i, task in enumerate(test_wo.task_list):
            if i != 1:
                task.complete(date.today(), str(g.people[0]), 1, 5 * random.random())
        test_wo.complete(date.today())

        excelname: str = test_wo.get_excel_filepath().split("\\")[-1]

        self.assertTrue(os.path.isfile(test_wo.get_excel_filepath()))
        self.assertTrue(os.path.isfile(test_wo.get_twois_filepath()))
        cmpdir: str = os.path.join(g.COMPLETE_DIR, str(date.today().year))
        self.assertFalse(os.path.isdir(cmpdir))
        cmpdir = os.path.join(cmpdir, get_date_dir_name(date.today()))
        self.assertFalse(os.path.isdir(cmpdir))
        self.assertFalse(os.path.isfile(os.path.join(cmpdir, excelname)))
        self.assertFalse(os.path.isfile(os.path.join(cmpdir,f"{test_wo.get_full_workorder_number()}_description.txt")))



class WorkOrderSmallerTests(unittest.TestCase):
    def test_extract_pending_number_works(self) -> None:
        result: int = extract_pending_twois_number("Pending-004")
        self.assertEqual(result, 4)


    def test_first_pending_wo_number_is_good(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        test_wo = WorkOrder()
        result: str = test_wo.wo_number
        self.assertEqual(result, "Pending-001")


    def test_second_pending_wo_number_is_good(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        wo = WorkOrder()
        wo.save()
        test_wo = WorkOrder()
        result :str = test_wo.wo_number
        self.assertEqual(result, "Pending-002")
        clear_dir(g.IN_PROGRESS_DIR)
        

    def test_fifth_pending_wo_number_is_good(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        for i in range(4):
            wo = WorkOrder()
            wo.save()
        test_wo = WorkOrder()
        result :str = test_wo.wo_number
        self.assertEqual(result, "Pending-005")


    def test_middle_pending_wo_number_is_good(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        for i in range(8):
            wo = WorkOrder()
            wo.save()
        rmfile :str = os.path.join(g.IN_PROGRESS_DIR, "Pending-005.twois")
        os.remove(rmfile)
        wo = WorkOrder()
        result :str = wo.wo_number
        self.assertEqual(result, "Pending-005")

    
    def test_is_valid_work_order1(self) -> None:
        self.assertTrue(is_a_valid_wo_number("123456VBS"))
        
    
    def test_is_valid_work_order2(self) -> None:
        self.assertTrue(is_a_valid_wo_number("123456"))
        
    
    def test_is_valid_work_order3(self) -> None:
        self.assertTrue(is_a_valid_wo_number("123456FGI"))
        
    
    def test_is_valid_work_order4(self) -> None:
        self.assertTrue(is_a_valid_wo_number("123456I"))
        
    
    def test_workorder_number_is_too_long(self) -> None:
        self.assertFalse(is_a_valid_wo_number("1234567891"))
        
    
    def test_work_order_number_has_invalid_suffix(self) -> None:
        self.assertFalse(is_a_valid_wo_number("123456A"))
        
    # BCOBB: matches_file() tests
    # BCOBB: is_approved() tests
    # BCOBB: get_full_workorder_number() tests
    
    
    def test_filematch__file_matches_itself_unapproved(self) -> None:
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(title=titlestr, due_date=date.today())
        test_wo.save()
        self.assertTrue(test_wo.matches_file(test_wo.get_excel_filepath()))
    
    
    def test_filematch__file_matches_itself_approved(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(wo_number="123456", title=titlestr, due_date=date.today())
        test_wo.save()
        wb: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws: Worksheet = wb.active
        insert_approval_stamps(ws)
        wb.save(test_wo.get_excel_filepath())
        wb.close()
        self.assertEqual(approved_file_status(test_wo.get_excel_filepath()), ExcelFileStatus.IS_VALID)
        self.assertTrue(test_wo.matches_file(test_wo.get_excel_filepath()))
    
    
    def test_filematch__file_mismatch_if_wrong_title(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(title=titlestr, due_date=date.today())
        test_wo.save()
        wb: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws: Worksheet = wb.active
        ws['D7'].value = titlestr[:-1]
        wb.save(test_wo.get_excel_filepath())
        wb.close()
        self.assertFalse(test_wo.matches_file(test_wo.get_excel_filepath()))
        ##BCOBB
    
    
    def test_filematch__file_mismatch_if_wrong_date(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(title=titlestr, due_date=date(2024,8,5))
        test_wo.save()
        wb: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws: Worksheet = wb.active
        ws['B3'].value = date_to_string(date.today())
        wb.save(test_wo.get_excel_filepath())
        wb.close()
        self.assertFalse(test_wo.matches_file(test_wo.get_excel_filepath()))
    
    
    def test_filematch__file_mismatch_if_wrong_tasklist(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(title=titlestr, due_date=date(2024,8,5),
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        test_wo.save()
        wb: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws: Worksheet = wb.active
        ws['B17'].value = "PPS-AV"
        wb.save(test_wo.get_excel_filepath())
        wb.close()
        self.assertFalse(test_wo.matches_file(test_wo.get_excel_filepath()))
    
    
    def test_filematch__file_mismatch_if_wrong_wonumber(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(wo_number="123456", title=titlestr, due_date=date(2024,8,5),
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        test_wo.save()
        wb: Workbook = load_workbook(test_wo.get_excel_filepath())
        ws: Worksheet = wb.active
        ws['A7'].value = "234567"
        insert_approval_stamps(ws)
        wb.save(test_wo.get_excel_filepath())
        wb.close()
        self.assertFalse(test_wo.matches_file(test_wo.get_excel_filepath()))
    
    
    def test_filematch__file_mismatch_if_invalid_file(self) -> None:
        clear_dir(g.IN_PROGRESS_DIR)
        titlestr: str = "14-Day Antivirus Updates (Ops GMM)"
        test_wo: WorkOrder = WorkOrder(wo_number="123456", title=titlestr, due_date=date(2024,8,5),
                                       task_list=[TaskItem(0, "Safety Message", "", 15),
                                                  TaskItem(10, "PPS-AV", "D743-25989-1", 16),
                                                  TaskItem(20, "PPS-AV 2", "D743-25989-2", 17)])
        self.assertFalse(test_wo.matches_file('invalid_file1.xlsx'))
        ##BCOBB

if __name__ == '__main__':
    unittest.main()
