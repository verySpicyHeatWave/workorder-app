"""This test file is meant to ensure that a workorder object being generated from an Excel
spreadsheet will correctly parse and populate its fields based on the contents of that
spreadsheet.
"""

#pylint: skip-file

import os
import unittest
from datetime import date, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from appfiles.library.logcomment import LogComment
from appfiles.library.site import Site
from appfiles.library.special import Special
from appfiles.library.taskitem import TaskItem
from appfiles.library.workorder_type import WorkOrderType
from appfiles.library.workorder import WorkOrder
from appfiles.utils.appglobals import IN_PROGRESS_DIR
from appfiles.utils.utils_for_testing import clear_dir, write_to_single_cell_in_test_file
from appfiles.utils.utils_for_testing import testfile



class WorkOrderExcelDateTests(unittest.TestCase):
    """For testing the due_date field of the work order class"""
    def test_due_date__is_a_date(self) -> None:
        """Checks to see if the due_date as pulled from the excel spreadsheet is a date object."""
        test_date: date = date(2023,8,16)
        write_to_single_cell_in_test_file(test_date.strftime("%#m/%#d/%Y"), 'B3')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(isinstance(test_wo.due_date, date))


    def test_due_date__is_correct(self) -> None:
        """Checks to see if the due_date as pulled from the excel spreadsheet is correct."""
        test_date: date = date(2023,8,16)
        write_to_single_cell_in_test_file(test_date.strftime("%#m/%#d/%Y"), 'B3')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.due_date, test_date)


    def test_due_date__is_today_if_blank(self) -> None:
        """Checks to see if the due_date defaults to today's date if left blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.due_date, date.today())


    def test_due_date__is_today_if_bad_string(self) -> None:
        """Checks to see if the due_date defaults to today's date if there's a non-date string."""
        write_to_single_cell_in_test_file('pickle', 'B3')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.due_date, date.today())




class WorkOrderExcelWONumberTests(unittest.TestCase):
    """For testing the work order number field of the work order class"""
    def test_wo_number__is_correct_1(self) -> None:
        """Checks to see if the wo_number as pulled from the excel spreadsheet is correct,
        along with the default suffix.
        """
        wo_str: str = "123456"
        write_to_single_cell_in_test_file(wo_str, 'A7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.get_full_workorder_number(), f"{wo_str}VBS")


    def test_wo_number__is_correct_2(self) -> None:
        """Checks to see if the wo_number as pulled from the excel spreadsheet is correct,
        along with the default suffix.
        """
        wo_str: str = "999999"
        write_to_single_cell_in_test_file(wo_str, 'A7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.get_full_workorder_number(), f"{wo_str}VBS")


    def test_wo_number__is_correct_3(self) -> None:
        """Checks to see if the wo_number as pulled from the excel spreadsheet is correct,
        along with the intentional suffix.
        """
        clear_dir(IN_PROGRESS_DIR)
        wo_str: str = "246810"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['A7'].value = wo_str
        ws['B7'].value = "FG"
        ws['C7'].value = "I"
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.get_full_workorder_number(), f"{wo_str}FGI")


    def test_wo_number__is_correct_4(self) -> None:
        """Checks to see if the wo_number as pulled from the excel spreadsheet is correct,
        along with the intentional suffix.
        """
        clear_dir(IN_PROGRESS_DIR)
        wo_str: str = "867539"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['A7'].value = wo_str
        ws['B7'].value = "EA"
        ws['C7'].value = "N"
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.get_full_workorder_number(), f"{wo_str}EAN")


    def test_wo_number__is_pending_if_blank(self) -> None:
        """Checks to see if the wo_number is set to a pending number if left blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.get_full_workorder_number(), "Pending-001")


    def test_wo_number__is_pending_if_bad_string(self) -> None:
        """Checks to see if the wo_number is set to a pending number if it contains
        a non-work-order-number string.
        """
        write_to_single_cell_in_test_file('pickle', 'A7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.get_full_workorder_number(), "Pending-001")




class WorkOrderExcelEnumTests(unittest.TestCase):
    """For testing the various enums associated with the work order class."""
    def test_site__parses_correctly1(self) -> None:
        """Checks to see if the site enum parses correctly from the string in the excel sheet."""
        write_to_single_cell_in_test_file('fg', 'B7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.site, Site.FG)


    def test_site__parses_correctly2(self) -> None:
        """Checks to see if the site enum parses correctly from the string in the excel sheet."""
        write_to_single_cell_in_test_file('eA', 'B7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.site, Site.EA)


    def test_site__parses_correctly3(self) -> None:
        """Checks to see if the site enum parses correctly from the string in the excel sheet."""
        write_to_single_cell_in_test_file('FD', 'B7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.site, Site.FD)


    def test_site__defaults_if_blank(self) -> None:
        """Checks to see if the site enum defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.site, Site.VB)


    def test_site__defaults_if_bad_string(self) -> None:
        """Checks to see if the site enum defaults correctly when the cell has a bad value."""
        write_to_single_cell_in_test_file('pickle', 'B7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.site, Site.VB)


    def test_special__parses_correctly(self) -> None:
        """Checks to see if the special enum parses correctly from the string in the
        excel sheet"""
        write_to_single_cell_in_test_file('I', 'C7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.special, Special.I)


    def test_special__defaults_if_blank(self) -> None:
        """Checks to see if the special enum defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.special, Special.S)


    def test_special__defaults_if_bad_string(self) -> None:
        """Checks to see if the special enum defaults correctly when the cell has a bad value."""
        write_to_single_cell_in_test_file('pickle', 'C7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.special, Special.S)


    def test_type__parses_correctly(self) -> None:
        """Checks to see if the work order type enum parses correctly from the string in the
        excel sheet.
        """
        write_to_single_cell_in_test_file('CM', 'I7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.type, WorkOrderType.CM)


    def test_type__defaults_if_blank(self) -> None:
        """Checks to see if the work order type enum defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.type, WorkOrderType.OTH)


    def test_type__defaults_if_bad_string(self) -> None:
        """Checks to see if the work order type enum defaults correctly when the cell
        has a bad value.
        """
        write_to_single_cell_in_test_file('pickle', 'I7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.type, WorkOrderType.OTH)




class WorkOrderExcelOtherFieldTests(unittest.TestCase):
    """Tests the fields which are a little less critical and/or easier to test"""
    def test_title__populates_correctly(self) -> None:
        """Checks to see if the title value populates correctly"""
        t: str = "14 Day AV Updates (OPS GMM)"
        write_to_single_cell_in_test_file(t, 'D7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.title, t)


    def test_title__truncates_if_too_long(self) -> None:
        """Checks to see if the title will be truncated to 50 characters if the cell contents
        are too long."""
        t: str = "Lorem ipsum dolor sit amet consectetur adipiscing elit "
        t += "sed do eiusmod tempor incididunt ut labore et dolore magna aliqua"
        title_limit: str = t[:WorkOrder.MAX_TITLE]
        write_to_single_cell_in_test_file(t, 'D7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.title, title_limit)


    def test_title__cleans_up_for_file_path(self) -> None:
        """Checks to see if the title will modify the title and remove characters that shouldn't
        be in a system filepath."""
        t: str = "This is a f%$#@<>! crazy string wow"
        t_fixed: str = "This is a f crazy string wow"
        write_to_single_cell_in_test_file(t, 'D7')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.title, t_fixed)


    def test_priority__populates_correctly(self) -> None:
        """Checks to see if the priority value populates correctly"""
        write_to_single_cell_in_test_file('1', 'A10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.priority, 1)


    def test_priority__defaults_if_blank(self) -> None:
        """Checks to see if the priority value defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.priority, 3)


    def test_priority__defaults_if_bad_value(self) -> None:
        """Checks to see if the priority value defaults correctly when the cell has a
        bad value."""
        write_to_single_cell_in_test_file('5', 'A10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.priority, 3)


    def test_priority__defaults_if_bad_string(self) -> None:
        """Checks to see if the priority value defaults correctly when the cell has 
        bad string contents."""
        write_to_single_cell_in_test_file('pickle', 'A10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.priority, 3)


    def test_creator__populates_correctly1(self) -> None:
        """Checks to see if the creator name string populates correctly"""
        creator_name: str = "Joel DeLeon"
        write_to_single_cell_in_test_file(creator_name, 'B10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.creator, creator_name)


    def test_creator__populates_correctly2(self) -> None:
        """Checks to see if the creator name string populates correctly"""
        creator_name: str = "Brian Cobb / Doug Orona"
        write_to_single_cell_in_test_file(creator_name, 'B10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.creator, creator_name)


    def test_creator__populates_correctly3(self) -> None:
        """Checks to see if the creator name string populates correctly"""
        creator_name: str = "Fu Manchu and an Elephant of Piss"
        write_to_single_cell_in_test_file(creator_name, 'B10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.creator, creator_name)


    def test_creator__defaults_if_blank(self) -> None:
        """Checks to see if the creator name defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.creator, "Brian Cobb")


    def test_related_wo__populates_correctly(self) -> None:
        """Checks to see if the related work order name string populates correctly"""
        wo_number: str = "415246I"
        write_to_single_cell_in_test_file(wo_number, 'G10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.related_wo, wo_number)


    def test_related_wo__na_if_blank(self) -> None:
        """Checks to see if the related work order name string populates correctly"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.related_wo, 'N/A')


    def test_related_wo__na_if_bad_string(self) -> None:
        """Checks to see if the related work order name string populates correctly"""
        wo_number: str = "i'm dirt"
        write_to_single_cell_in_test_file(wo_number, 'G10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.related_wo, 'N/A')


    def test_ncr_number__na_if_ncr_not_required1(self) -> None:
        """Checks whether the NCR number string is 'N/A' if an NCR isn't required"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.ncr_number, 'N/A')


    def test_ncr_number__na_if_ncr_not_required2(self) -> None:
        """Checks whether the NCR number string is 'N/A' if an NCR isn't required, even
        if I feed it an NCR number"""
        ncr_number = "NCR123456W"
        write_to_single_cell_in_test_file(ncr_number, 'B12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.ncr_number, 'N/A')


    def test_ncr_number__na_if_ncr_not_required3(self) -> None:
        """Checks whether the NCR number string is 'N/A' if an NCR isn't required, even
        if I feed it a non-valid string"""
        ncr_number = "THIS PLACE SUCKS"
        write_to_single_cell_in_test_file(ncr_number, 'B12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.ncr_number, 'N/A')


    def test_ncr_number__populates_if_ncr_required1(self) -> None:
        """Checks whether the NCR number string populates correctly if ncr_required and if
        the NCR string is valid"""
        clear_dir(IN_PROGRESS_DIR)
        ncr_number = "NCR123456W"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['B12'].value = ncr_number
        ws['A12'].value = 'YES'
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.ncr_number, ncr_number)


    def test_ncr_number__complains_if_blank_and_ncr_required(self) -> None:
        """Checks whether the NCR number string is 'N/A' if an NCR isn't required, even
        if I feed it an NCR number"""
        write_to_single_cell_in_test_file('yes', 'A12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.ncr_number, 'REQUIRED')


    def test_ncr_number__complains_if_bad_string_and_ncr_required(self) -> None:
        """Checks whether the NCR number string is 'N/A' if an NCR isn't required, even
        if I feed it a non-valid string"""
        clear_dir(IN_PROGRESS_DIR)
        ncr_number = "THIS PLACE SUCKS"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['B12'].value = ncr_number
        ws['A12'].value = 'YES'
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.ncr_number, 'REQUIRED')




class WorkOrderExcelBldgAndRoomTests(unittest.TestCase):
    """Tests how well the room and building values populate"""
    def test_building__populates_correctly1(self) -> None:
        """Checks to see if the building int parses and populates correctly from cell D8."""
        bldg: str = "bldg. 5678"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 5678)


    def test_building__populates_correctly2(self) -> None:
        """Checks to see if the building int parses and populates correctly from cell F8."""
        bldg: str = "b5678"
        write_to_single_cell_in_test_file(bldg, 'F8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 5678)


    def test_building__populates_correctly3(self) -> None:
        """Checks to see if the building int parses and populates correctly from cell H8."""
        bldg: str = "b.5678"
        write_to_single_cell_in_test_file(bldg, 'H8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 5678)


    def test_building__populates_correctly4(self) -> None:
        """Checks to see if the building int parses and populates correctly from cell J8."""
        bldg: str = "B 5678"
        write_to_single_cell_in_test_file(bldg, 'J8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 5678)


    def test_building__populates_correctly5(self) -> None:
        """Checks to see if the building int parses and populates correctly
        with the room number.
        """
        bldg: str = "bldg 1945, room 111"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)


    def test_building__defaults_if_blank(self) -> None:
        """Checks to see if the building int defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1768)


    def test_building__defaults_if_bad_string(self) -> None:
        """Checks to see if the building int defaults correctly when the cell has 
        bad string contents."""
        bldg: str = "you can't explain that!"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1768)


    def test_building__lower_clamp1(self) -> None:
        """Checks to see if the building int defaults if the value is below the lower clamp."""
        bldg: str = "b9"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1768)


    def test_building__lower_clamp2(self) -> None:
        """Checks to see if the building int takes right at the lower clamp."""
        bldg: str = "b10"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 10)


    def test_building__higher_clamp1(self) -> None:
        """Checks to see if the building int defaults if the value is above the higher clamp."""
        bldg: str = "b100000"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1768)


    def test_building__higher_clamp2(self) -> None:
        """Checks to see if the building int takes right at the higher clamp."""
        bldg: str = "b99999"
        write_to_single_cell_in_test_file(bldg, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 99_999)


    def test_room__populates_correctly1(self) -> None:
        """Checks to see if the room int parses and populates correctly from cell F8."""
        rm: str = "rm 111"
        write_to_single_cell_in_test_file(rm, 'F8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 111)


    def test_room__populates_correctly2(self) -> None:
        """Checks to see if the room int parses and populates correctly from cell H8."""
        rm: str = "rm.23"
        write_to_single_cell_in_test_file(rm, 'H8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 23)


    def test_room__populates_correctly3(self) -> None:
        """Checks to see if the room int parses and populates correctly from cell J8."""
        rm: str = "r6"
        write_to_single_cell_in_test_file(rm, 'J8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__populates_correctly4(self) -> None:
        """Checks to see if the room int parses and populates correctly from cell D8."""
        rm: str = "r 99"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 99)


    def test_room__populates_correctly5(self) -> None:
        """Checks to see if the room int parses and populates correctly with the
        'room 6/7/8' string."""
        rm: str = "room 6/7/8"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__populates_correctly6(self) -> None:
        """Checks to see if the room int parses and populates correctly with the
        'room 6,7,8' string."""
        rm: str = "room 6,7,8"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__defaults_if_blank(self) -> None:
        """Checks to see if the room int defaults correctly when the cell is blank."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__defaults_if_bad_string(self) -> None:
        """Checks to see if the room int defaults correctly when the cell has 
        bad string contents."""
        rm: str = "WHAT IN THE DICKENS"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__lower_clamp1(self) -> None:
        """Checks to see if the room int defaults if the value is below the lower clamp."""
        rm: str = "r0"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__lower_clamp2(self) -> None:
        """Checks to see if the room int takes right at the lower clamp."""
        rm: str = "r1"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 1)


    def test_room__higher_clamp1(self) -> None:
        """Checks to see if the room int defaults if the value is above the higher clamp."""
        rm: str = "rm1000"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 6)


    def test_room__higher_clamp2(self) -> None:
        """Checks to see if the room int takes right at the higher clamp."""
        rm: str = "rm999"
        write_to_single_cell_in_test_file(rm, 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.room, 999)


    def test_building_and_room__populate_from_any_cell1(self) -> None:
        """Checks to see if the room int parses and populates correctly whe they're both
        in a sort of random cell."""
        clear_dir(IN_PROGRESS_DIR)
        bldg: str = "bldg. 1945"
        room: str = "room 111"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['F8'].value = bldg
        ws['D8'].value = room
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)
        self.assertEqual(test_wo.room, 111)


    def test_building_and_room__populate_from_any_cell2(self) -> None:
        """Checks to see if the room int parses and populates correctly whe they're both
        in a sort of random cell."""
        clear_dir(IN_PROGRESS_DIR)
        bldg: str = "bldg. 1945"
        room: str = "room 111"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['D8'].value = bldg
        ws['J8'].value = room
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)
        self.assertEqual(test_wo.room, 111)


    def test_building_and_room__populate_from_any_cell3(self) -> None:
        """Checks to see if the room int parses and populates correctly whe they're both
        in a sort of random cell."""
        clear_dir(IN_PROGRESS_DIR)
        bldg: str = "bldg. 1945"
        room: str = "room 111"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['H8'].value = bldg
        ws['D8'].value = room
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)
        self.assertEqual(test_wo.room, 111)


    def test_building_and_room__populate_from_any_cell4(self) -> None:
        """Checks to see if the room int parses and populates correctly whe they're both
        in a sort of random cell."""
        bldg: str = "bldg. 1945"
        room: str = "room 111"
        write_to_single_cell_in_test_file(f"{room} {bldg}", 'D8')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)
        self.assertEqual(test_wo.room, 111)


    def test_building_and_room__populate_from_any_cell5(self) -> None:
        """Checks to see if the room int parses and populates correctly whe they're both
        in a sort of random cell."""
        clear_dir(IN_PROGRESS_DIR)
        bldg: str = "bldg. 1945"
        room: str = "room 111"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['H8'].value = bldg
        ws['J8'].value = room
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)
        self.assertEqual(test_wo.room, 111)


    def test_building_and_room__populate_from_any_cell6(self) -> None:
        """Checks to see if the room int parses and populates correctly whe they're both
        in a sort of random cell."""
        clear_dir(IN_PROGRESS_DIR)
        bldg: str = "bldg. 1945"
        room: str = "room 111"
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        ws['F8'].value = bldg
        ws['H8'].value = room
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.building, 1945)
        self.assertEqual(test_wo.room, 111)




class WorkOrderExcelBooleanFieldTests(unittest.TestCase):
    """For testing whether all of those boolean values will populate correctly."""
    def test_pac_required__populates_correctly_false(self) -> None:
        """Checks to see if PAC Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file("PAC - NO", 'I10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.pac_required)


    def test_pac_required__populates_correctly_true1(self) -> None:
        """Checks to see if PAC Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file("PAC - yes", 'I10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.pac_required)


    def test_pac_required__populates_correctly_true2(self) -> None:
        """Checks to see if PAC Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file("PAC - REQUIRED", 'I10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.pac_required)


    def test_pac_required__defaults_to_false_if_blank(self) -> None:
        """Checks to see if PAC Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.pac_required)


    def test_pac_required__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if PAC Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'I10')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.pac_required)


    def test_ncr_required__populates_correctly_false(self) -> None:
        """Checks to see if NCR Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'A12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.ncr_required)


    def test_ncr_required__populates_correctly_true(self) -> None:
        """Checks to see if NCR Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'A12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.ncr_required)


    def test_ncr_required__defaults_to_false_if_blank(self) -> None:
        """Checks to see if NCR Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.ncr_required)


    def test_ncr_required__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if NCR Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'A12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.ncr_required)


    def test_task_lead_required__populates_correctly_false(self) -> None:
        """Checks to see if Task Lead Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'E12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.task_lead_required)


    def test_task_lead_required__populates_correctly_true(self) -> None:
        """Checks to see if Task Lead Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'E12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.task_lead_required)


    def test_task_lead_required__defaults_to_false_if_blank(self) -> None:
        """Checks to see if Task Lead Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.task_lead_required)


    def test_task_lead_required__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if Task Lead Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'E12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.task_lead_required)


    def test_tech_witness_point__populates_correctly_false(self) -> None:
        """Checks to see if Tech Witness Point Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'F12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.tech_witness_point)


    def test_tech_witness_point__populates_correctly_true(self) -> None:
        """Checks to see if Tech Witness Point Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'F12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.tech_witness_point)


    def test_tech_witness_point__defaults_to_false_if_blank(self) -> None:
        """Checks to see if Tech Witness Point Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.tech_witness_point)


    def test_tech_witness_point__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if Tech Witness Point Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'F12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.tech_witness_point)


    def test_peer_review_required__populates_correctly_false(self) -> None:
        """Checks to see if Peer Review Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'G12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.peer_review_required)


    def test_peer_review_required__populates_correctly_true(self) -> None:
        """Checks to see if Peer Review Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'G12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.peer_review_required)


    def test_peer_review_required__defaults_to_false_if_blank(self) -> None:
        """Checks to see if Peer Review Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.peer_review_required)


    def test_peer_review_required__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if Peer Review Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'G12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.peer_review_required)


    def test_peer_review_attached__populates_correctly_false(self) -> None:
        """Checks to see if Peer Review Attached boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'H12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.peer_review_attached)


    def test_peer_review_attached__populates_correctly_true(self) -> None:
        """Checks to see if Peer Review Attached boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'H12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.peer_review_attached)


    def test_peer_review_attached__defaults_to_false_if_blank(self) -> None:
        """Checks to see if Peer Review Attached boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.peer_review_attached)


    def test_peer_review_attached__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if Peer Review Attached boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'H12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.peer_review_attached)


    def test_ehs_required__populates_correctly_false(self) -> None:
        """Checks to see if EHS Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'I12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.ehs_required)


    def test_ehs_required__populates_correctly_true(self) -> None:
        """Checks to see if EHS Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'I12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.ehs_required)


    def test_ehs_required__defaults_to_false_if_blank(self) -> None:
        """Checks to see if EHS Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.ehs_required)


    def test_ehs_required__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if EHS Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'I12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.ehs_required)


    def test_qamip__populates_correctly_false(self) -> None:
        """Checks to see if QAMIP Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'J12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.qamip)


    def test_qamip__populates_correctly_true(self) -> None:
        """Checks to see if QAMIP Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'J12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.qamip)


    def test_qamip__defaults_to_false_if_blank(self) -> None:
        """Checks to see if QAMIP Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.qamip)


    def test_qamip__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if QAMIP Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'J12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.qamip)


    def test_qa_review_required__populates_correctly_false(self) -> None:
        """Checks to see if QA Review Required boolean populates correctly to false"""
        write_to_single_cell_in_test_file('No', 'K12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.qa_review_required)


    def test_qa_review_required__populates_correctly_true(self) -> None:
        """Checks to see if QA Review Required boolean populates correctly to true"""
        write_to_single_cell_in_test_file('Yes', 'K12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertTrue(test_wo.qa_review_required)


    def test_qa_review_required__defaults_to_false_if_blank(self) -> None:
        """Checks to see if QA Review Required boolean defaults to FALSE"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.qa_review_required)


    def test_qa_review_required__defaults_to_false_if_bad_string(self) -> None:
        """Checks to see if QA Review Required boolean defaults to FALSE"""
        write_to_single_cell_in_test_file("BAAAAAD companyyyyy", 'K12')
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertFalse(test_wo.qa_review_required)




class WorkOrderExcelTaskItemTests(unittest.TestCase):
    """For testing whether or not all of those boolean values will populate correctly."""
    def test_task_list__populates_correctly1(self) -> None:
        """Checks to see that the task list populates as expected with good values"""
        clear_dir(IN_PROGRESS_DIR)
        tasks: list[TaskItem] = [TaskItem(0, "Safety Message", "", 15)]
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        for i in range(1, 4):
            taskno: int = i * 10
            row: int = i + 15
            summary: str = f"This is task number {i}"
            reference: str = f"This is reference number {i}"
            ws[f"A{row}"].value = taskno
            ws[f"B{row}"].value = summary
            ws[f"G{row}"].value = reference
            tasks.append(TaskItem(taskno, summary, reference, row))
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        for i, task in enumerate(tasks):
            self.assertEqual(task, test_wo.task_list[i])


    def test_task_list__populates_correctly2(self) -> None:
        """Checks to see that the task list populates as expected with good values,
        even when the entries are a little spaced out"""
        clear_dir(IN_PROGRESS_DIR)
        tasks: list[TaskItem] = [TaskItem(0, "Safety Message", "", 15)]
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        for i in range(1, 4):
            taskno: int = i * 10
            row: int = i + i + 15
            summary: str = f"This is task number {i}"
            reference: str = f"This is reference number {i}"
            ws[f"A{row}"].value = taskno
            ws[f"B{row}"].value = summary
            ws[f"G{row}"].value = reference
            tasks.append(TaskItem(taskno, summary, reference, row))
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        for i, task in enumerate(tasks):
            self.assertEqual(task, test_wo.task_list[i])


    def test_task_list__populates_correctly3(self) -> None:
        """Checks to see that the task list populates as expected with good values,
        even when the task numbers are not multiples of ten"""
        clear_dir(IN_PROGRESS_DIR)
        tasks: list[TaskItem] = [TaskItem(0, "Safety Message", "", 15)]
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        for i in range(1, 4):
            taskno: int = i * 12
            row: int = i + 15
            summary: str = f"This is task number {i}"
            reference: str = f"This is reference number {i}"
            ws[f"A{row}"].value = taskno
            ws[f"B{row}"].value = summary
            ws[f"G{row}"].value = reference
            tasks.append(TaskItem(taskno, summary, reference, row))
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        for i, task in enumerate(tasks):
            self.assertEqual(task, test_wo.task_list[i])


    def test_task_list_length_is_two_with_only_one_task(self) -> None:
        """Checks to see that the task list gets a default item complaining about a lack of tasks
        if at least one additional task is not provided."""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(len(test_wo.task_list), 2)


    def test_task_list_default_second_task(self) -> None:
        """Checks to see that the task list gets a default item complaining about a lack of tasks
        if at least one additional task is not provided."""
        clear_dir(IN_PROGRESS_DIR)
        default_task: TaskItem = TaskItem(10, "TASK DESCRIPTION REQUIRED",
                                          "REFERENCE REQUIRED", 16)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(test_wo.task_list[1], default_task)


    def test_task_list__complains_about_missing_reference(self) -> None:
        """Checks to see that the task list complains about a missing reference if one of
        the task items is missing a reference string"""
        clear_dir(IN_PROGRESS_DIR)
        tasks: list[TaskItem] = [TaskItem(0, "Safety Message", "", 15)]
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        for i in range(1, 3):
            taskno: int = i * 12
            row: int = i + 15
            summary: str = f"This is task number {i}"
            reference: str = f"This is reference number {i}"
            ws[f"A{row}"].value = taskno
            ws[f"B{row}"].value = summary
            ws[f"G{row}"].value = reference if i != 2 else ""
            tasks.append(TaskItem(taskno, summary, reference, row))
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual("REFERENCE REQUIRED", test_wo.task_list[2].reference)



class WorkOrderExcelSheetLoadingTests(unittest.TestCase):
    """For testing whether or not all of those boolean values will populate correctly."""
    def is_a_default_workorder(self, file: str) -> bool:
        """Method for this class which determines if a work order had been generated
        with default settings
        """
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(file)
        resp: bool = (test_wo.building == 1768 and
                      test_wo.creator == "Brian Cobb" and
                      test_wo.title == "No title provided")
        return resp


    def test__values_are_default_if_is_bad_filepath(self) -> None:
        """Checks to see that values tend towards defaults when a non-existent file is loaded"""
        self.assertTrue(self.is_a_default_workorder('this_is_not_a_file.txt'))


    def test__values_are_default_if_is_invalid_file(self) -> None:
        """Checks to see that values tend towards defaults when a non-xlsx file is loaded"""
        fname: str = "invalid_file.txt"
        with open(fname, "w", encoding="utf-8") as f:
            f.write("This is not an excel file!")
        self.assertTrue(os.path.isfile(fname)) # make sure the file actually exists
        self.assertTrue(self.is_a_default_workorder(fname))


    def test__values_are_default_if_is_invalid_xlsx1(self) -> None:
        """Checks to see that values tend towards defaults when a non-twois xlsx is loaded"""
        clear_dir(IN_PROGRESS_DIR)
        fname: str = "invalid_file1.xlsx"
        self.assertTrue(os.path.isfile(fname)) # make sure the file actually exists
        self.assertTrue(self.is_a_default_workorder(fname))


    def test__values_are_default_if_is_invalid_xlsx2(self) -> None:
        """Checks to see that values tend towards defaults when a non-twois xlsx is loaded"""
        clear_dir(IN_PROGRESS_DIR)
        fname: str = "invalid_file3.xlsx"
        self.assertTrue(os.path.isfile(fname)) # make sure the file actually exists
        self.assertTrue(self.is_a_default_workorder(fname))

class WorkOrderLogCommentsTests(unittest.TestCase):
    """For testing whether or not the log comments list will populate correctly."""
    def test_comment_list__populates_correctly1(self) -> None:
        """Checks to see that the comment list populates as expected with good values"""
        clear_dir(IN_PROGRESS_DIR)
        comments: list[LogComment] = []
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        for i in range(4):
            row = i + 86
            name: str = "B. Cobb / 3677837"
            text: str = f"This is comment number {i + 1}"
            c_date: date = date.today() + timedelta(days=i)
            ws[f"A{row}"].value = text
            ws[f"I{row}"].value = name
            ws[f"K{row}"].value = c_date.strftime("%#m/%#d/%Y")
            comments.append(LogComment(text, name, c_date, i))
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        for i, comment in enumerate(comments):
            self.assertEqual(comment, test_wo.comments[i])


    def test_comment_list__populates_correctly2(self) -> None:
        """Checks to see that the comment list populates as expected with good values,
        even when the entries are a little spaced out"""
        clear_dir(IN_PROGRESS_DIR)
        comments: list[LogComment] = []
        wb: Workbook = load_workbook(testfile)
        ws: Worksheet = wb.active
        for i in range(4):
            row = i + i + 86
            name: str = "B. Cobb / 3677837"
            text: str = f"This is comment number {i + 1}"
            c_date: date = date.today() + timedelta(days=i)
            ws[f"A{row}"].value = text
            ws[f"I{row}"].value = name
            ws[f"K{row}"].value = c_date.strftime("%#m/%#d/%Y")
            comments.append(LogComment(text, name, c_date, i))
        wb.save(testfile)
        wb.close()
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        for i, comment in enumerate(comments):
            self.assertEqual(comment, test_wo.comments[i])


    def test_comment_list__populates_correctly3(self) -> None:
        """Checks to see that the list is empty if there are no comments"""
        clear_dir(IN_PROGRESS_DIR)
        test_wo: WorkOrder = WorkOrder.from_xlsx(testfile)
        self.assertEqual(0, len(test_wo.comments))

if __name__ == '__main__':
    unittest.main()
